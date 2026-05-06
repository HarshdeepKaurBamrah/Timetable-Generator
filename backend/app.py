"""
app.py  –  Academic Timetable System (MongoDB edition)

Changes vs original:
  - SQLite -> MongoDB (pymongo)
  • Batches have custom names (e.g. "Batch Alpha", "Batch B1") supplied
    at division-creation time (or auto-generated as "Batch 1 … N")
  • GET /api/divisions returns batches with id, name, size, division_id
  • POST /api/divisions accepts batch_names[] list
  • GET /api/batches — new endpoint listing all batches with counts
"""
import json, os, sys, io, time, threading, uuid
from functools import wraps
from datetime import datetime
from flask import Flask, request, jsonify, session, send_file, send_from_directory, g
from pymongo.errors import BulkWriteError

sys.path.insert(0, os.path.dirname(__file__))
from constraints import constraints_ui_meta, normalize_constraint_document
from database import (
    active_timetable_filter,
    clear_active_timetable,
    get_db,
    get_active_timetable_version,
    get_startup_report,
    hash_pw,
    init_db,
    next_seq,
)
from solver import TimetableSolver, build_state_from_active_timetable, validate_active_timetable
from rescheduler import (
    move_slot, swap_slots, teacher_absent, claim_slot,
    change_teacher_for_slot, lock_slot, add_constraint, update_constraint,
    get_timetable, get_notifications, get_change_log, validate_or_apply_cell_move,
    validate_or_apply_swap
)

FRONTEND_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', 'frontend'))

app = Flask(__name__,
            static_folder=os.path.join(FRONTEND_DIR, 'static'),
            template_folder=FRONTEND_DIR)
app.secret_key = os.environ.get('SECRET_KEY', 'tt_mongo_secret_2024')
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'

GENERATION_JOBS = {}
GENERATION_LOCK = threading.Lock()
GENERATION_JOB_TTL_SECONDS = 3600
FAST_STRICT_SEARCH_LIMITS = {
    "labs": [8, 16],
    "lectures": [12, 28],
    "parallel": [8, 16],
}
FAST_STRICT_PHASE_SECONDS = 2.0

init_db()


def _sanitize_user(user):
    if not user:
        return None
    return {
        "id": str(user.get("int_id")),
        "int_id": int(user.get("int_id")),
        "name": user.get("name"),
        "role": user.get("role"),
        "teacher_id": user.get("teacher_id"),
        "email": user.get("email"),
    }


def _auth_token_from_request():
    auth_header = str(request.headers.get("Authorization") or "").strip()
    if not auth_header.lower().startswith("bearer "):
        return None
    token = auth_header[7:].strip()
    return token or None


def _issue_auth_token(user):
    db = get_db()
    token = uuid.uuid4().hex
    db.auth_tokens.insert_one(
        {
            "token": token,
            "user_int_id": int(user["int_id"]),
            "created_at": datetime.now().isoformat(),
        }
    )
    return token


def _revoke_auth_token(token):
    if not token:
        return
    get_db().auth_tokens.delete_many({"token": token})


def _load_user_from_token(token):
    if not token:
        return None
    db = get_db()
    token_row = db.auth_tokens.find_one({"token": token}, {"_id": 0})
    if not token_row:
        return None
    user = db.users.find_one({"int_id": int(token_row["user_int_id"])}, {"_id": 0, "password": 0})
    return _sanitize_user(user)


def _current_user():
    if hasattr(g, "_current_user_loaded"):
        return getattr(g, "current_user", None)
    token = _auth_token_from_request()
    g._current_user_loaded = True
    g.current_auth_token = token
    if token:
        g.current_user = _load_user_from_token(token)
        return g.current_user
    if "user_id" not in session:
        g.current_user = None
        return None
    g.current_user = {
        "id": str(session["user_id"]),
        "int_id": int(session["user_id"]),
        "name": session.get("name"),
        "role": session.get("role"),
        "teacher_id": session.get("teacher_id"),
        "email": session.get("email"),
    }
    return g.current_user


def _current_user_value(key, default=None):
    user = _current_user()
    if not user:
        return default
    return user.get(key, default)


def _notification_visibility_filter():
    role = _current_user_value("role")
    teacher_id = _current_user_value("teacher_id")
    role_filter = {"$or": [{"for_role": role}, {"for_role": None}]}
    if teacher_id:
        return {"$and": [role_filter, {"$or": [{"for_teacher": teacher_id}, {"for_teacher": None}]}]}
    return role_filter


def _cleanup_generation_jobs():
    cutoff = time.time() - GENERATION_JOB_TTL_SECONDS
    with GENERATION_LOCK:
        stale = [
            job_id
            for job_id, job in GENERATION_JOBS.items()
            if job.get('finished_ts') and job['finished_ts'] < cutoff
        ]
        for job_id in stale:
            GENERATION_JOBS.pop(job_id, None)


def _generation_job_payload(job):
    if not job:
        return None
    return {
        "job_id": job["job_id"],
        "status": job.get("status", "queued"),
        "stage": job.get("stage"),
        "message": job.get("message", ""),
        "percent": job.get("percent", 0),
        "elapsed_seconds": round(float(job.get("elapsed_seconds", 0.0) or 0.0), 3),
        "started_at": job.get("started_at"),
        "updated_at": job.get("updated_at"),
        "finished_at": job.get("finished_at"),
        "details": job.get("details") or {},
        "diagnostic": job.get("diagnostic"),
        "diagnostics": job.get("diagnostics") or [],
        "coverage_report": job.get("coverage_report") or {},
        "version_id": job.get("version_id"),
        "slots": job.get("slots"),
        "lab_slots": job.get("lab_slots"),
        "partial": bool(job.get("partial")),
        "unscheduled": job.get("unscheduled") or [],
        "result": job.get("result"),
    }


def _create_generation_job(changed_by):
    _cleanup_generation_jobs()
    now = datetime.now().isoformat()
    job = {
        "job_id": uuid.uuid4().hex,
        "status": "queued",
        "stage": "queued",
        "message": "Generation request queued",
        "percent": 0,
        "details": {},
        "diagnostic": None,
        "diagnostics": [],
        "coverage_report": {},
        "result": None,
        "version_id": None,
        "slots": 0,
        "lab_slots": 0,
        "partial": False,
        "unscheduled": [],
        "changed_by": changed_by,
        "started_at": now,
        "updated_at": now,
        "finished_at": None,
        "finished_ts": None,
        "elapsed_seconds": 0.0,
    }
    with GENERATION_LOCK:
        GENERATION_JOBS[job["job_id"]] = job
    return _generation_job_payload(job)


def _update_generation_job(job_id, **updates):
    with GENERATION_LOCK:
        job = GENERATION_JOBS.get(job_id)
        if not job:
            return None
        job.update(updates)
        job["updated_at"] = datetime.now().isoformat()
        if job.get("status") in {"completed", "failed"}:
            if not job.get("finished_at"):
                job["finished_at"] = job["updated_at"]
            job["finished_ts"] = time.time()
        return _generation_job_payload(job)


def _get_generation_job(job_id):
    _cleanup_generation_jobs()
    with GENERATION_LOCK:
        return _generation_job_payload(GENERATION_JOBS.get(job_id))


def _get_active_generation_job():
    _cleanup_generation_jobs()
    with GENERATION_LOCK:
        active = [job for job in GENERATION_JOBS.values() if job.get("status") in {"queued", "running"}]
        if not active:
            return None
        active.sort(key=lambda job: job.get("started_at", ""), reverse=True)
        return _generation_job_payload(active[0])


def _diagnostic_dicts(items, limit=20):
    rows = []
    for item in items or []:
        if isinstance(item, dict):
            rows.append(dict(item))
        else:
            rows.append(dict(getattr(item, "__dict__", {}) or {}))
        if len(rows) >= limit:
            break
    return rows


def _write_generation_report(db, payload):
    try:
        doc = {
            "int_id": next_seq(db, "generation_reports"),
            "created_at": datetime.now().isoformat(),
            "status": payload.get("status", "completed"),
            "version_id": payload.get("version_id", get_active_timetable_version(db)),
            "message": payload.get("message") or payload.get("error") or "",
            "partial": bool(payload.get("partial")),
            "diagnostic": payload.get("diagnostic"),
            "diagnostics": payload.get("diagnostics") or [],
            "coverage_report": payload.get("coverage_report") or {},
            "unscheduled": payload.get("unscheduled") or [],
            "elapsed_seconds": payload.get("elapsed_seconds"),
            "slots": payload.get("slots", 0),
            "lab_slots": payload.get("lab_slots", 0),
        }
        db.generation_reports.insert_one(doc)
    except Exception as exc:
        app.logger.warning(f"Could not persist generation report: {exc}")


def _latest_generation_report(db):
    active_version = get_active_timetable_version(db)
    report = None
    if active_version is not None:
        report = db.generation_reports.find_one({"version_id": active_version}, {"_id": 0}, sort=[("created_at", -1)])
    if not report:
        report = db.generation_reports.find_one({}, {"_id": 0}, sort=[("created_at", -1)])
    return report or {}


def _suggestions_for_problem_codes(codes):
    suggestions = []
    code_set = set(codes or [])
    if code_set & {"teacher_missing", "subject_teacher_missing", "tutorial_teacher_missing", "lab_teacher_missing", "lab_teacher_invalid"}:
        suggestions.append("Assign the missing subject or batch teacher, then generate again.")
    if code_set & {"lecture_rooms_missing", "lab_rooms_missing", "room_missing", "room_overlap", "distinct_rooms_unavailable", "grouped_lab_room_shortage"}:
        suggestions.append("Add enough lecture/lab rooms or reduce simultaneous parallel/lab sessions.")
    if code_set & {"division_over_capacity"}:
        suggestions.append("Reduce weekly lecture/lab load for the division, or add more working periods/days.")
    if code_set & {"teacher_overlap", "teacher_daily_limit", "teacher_weekly_limit", "teacher_unavailable", "grouped_lab_teacher_overlap"}:
        suggestions.append("Increase teacher availability/limits or assign another teacher for the affected subject.")
    if code_set & {"avoid_day", "avoid_period", "once_per_day", "same_day_repeat", "max_per_day", "start_or_end_only", "end_only", "no_free_slots", "coverage_shortfall", "search_timeout", "dead_end"}:
        suggestions.append("Relax the hard constraint, add more periods/days, or drag-swap placed sessions to create room.")
    if code_set & {"lab_break_cross", "lab_duration", "lab_contiguity", "end_of_day_lab_coverage"}:
        suggestions.append("Keep labs in two consecutive periods, preferably the final valid two-period block.")
    if not suggestions and code_set:
        suggestions.append("Review the listed problem, adjust data or constraints, and run generation again.")
    return suggestions


def _build_timetable_problem_report(db):
    validation_diagnostics = [d.__dict__ for d in validate_active_timetable(db)]
    solver = TimetableSolver(
        period_ids=get_config('period_ids', ["P1","P2","P3","P4","P5","P6"]),
        days=get_config('days', ["Monday","Tuesday","Wednesday","Thursday","Friday"]),
        db=db,
    )
    solver.state = build_state_from_active_timetable(db, solver.context)
    end_lab_issue = solver.evaluator.validate_end_of_day_lab_coverage(solver.state)
    if end_lab_issue:
        validation_diagnostics.append(end_lab_issue.__dict__)
    _, coverage_report = solver.validate_required_coverage()
    coverage_diagnostics = [
        {
            "phase": "coverage",
            "code": "coverage_shortfall",
            "message": gap.get("message", "Required timetable coverage is incomplete"),
            "entities": gap,
        }
        for gap in (coverage_report.get("gaps") or [])
    ]
    latest = _latest_generation_report(db)
    unscheduled = latest.get("unscheduled") or []
    diagnostics = validation_diagnostics + coverage_diagnostics
    codes = [item.get("code") for item in diagnostics]
    codes.extend(item.get("reason_code") for item in unscheduled)
    return {
        "ok": not diagnostics and not unscheduled,
        "diagnostics": diagnostics,
        "validation_diagnostics": validation_diagnostics,
        "coverage_report": coverage_report,
        "unscheduled": unscheduled,
        "suggestions": _suggestions_for_problem_codes([code for code in codes if code]),
        "last_generation": latest,
    }


def _execute_generation(changed_by=None, progress_callback=None, allow_partial=False):
    db = get_db()
    period_ids = get_config('period_ids', ["P1","P2","P3","P4","P5","P6"])
    days = get_config('days', ["Monday","Tuesday","Wednesday","Thursday","Friday"])
    strict_search_limits = FAST_STRICT_SEARCH_LIMITS if allow_partial else None
    strict_phase_seconds = FAST_STRICT_PHASE_SECONDS if allow_partial else None
    solver = TimetableSolver(
        period_ids=period_ids,
        days=days,
        seed=42,
        db=db,
        progress_callback=progress_callback,
        search_limits=strict_search_limits,
        max_phase_seconds=strict_phase_seconds,
    )
    started_at = time.perf_counter()
    ok = solver.solve()
    elapsed_seconds = round(time.perf_counter() - started_at, 3)
    strict_failure = None
    if not ok:
        strict_failure = solver.failure or (solver.diagnostics[0] if solver.diagnostics else None)
        if not allow_partial:
            payload = {
                "error": strict_failure.message if strict_failure else "Could not generate timetable",
                "diagnostic": strict_failure.__dict__ if strict_failure else None,
                "diagnostics": _diagnostic_dicts(solver.diagnostics),
                "coverage_report": solver.coverage_report,
                "elapsed_seconds": elapsed_seconds,
                "partial": False,
                "unscheduled": [],
            }
            _write_generation_report(db, {**payload, "status": "failed"})
            return (payload, 400)
        partial_solver = TimetableSolver(period_ids=period_ids, days=days, seed=42, db=db, progress_callback=progress_callback)
        partial_ok = partial_solver.solve_partial()
        elapsed_seconds = round(time.perf_counter() - started_at, 3)
        if not partial_ok:
            partial_failure = partial_solver.failure or strict_failure or (partial_solver.diagnostics[0] if partial_solver.diagnostics else None)
            payload = {
                "error": partial_failure.message if partial_failure else "Could not generate even a partial timetable",
                "diagnostic": partial_failure.__dict__ if partial_failure else None,
                "diagnostics": _diagnostic_dicts(partial_solver.diagnostics or solver.diagnostics),
                "coverage_report": partial_solver.coverage_report or solver.coverage_report,
                "elapsed_seconds": elapsed_seconds,
                "partial": True,
                "unscheduled": partial_solver.unscheduled,
            }
            _write_generation_report(db, {**payload, "status": "failed"})
            return (payload, 400)
        solver = partial_solver

    if not solver.assignments and not solver.lab_assignments:
        payload = {
            "error": "No timetable slots generated. Check that subjects have weekly hours set and teachers assigned.",
            "diagnostic": None,
            "diagnostics": [],
            "coverage_report": solver.coverage_report,
            "elapsed_seconds": elapsed_seconds,
            "partial": bool(solver.partial),
            "unscheduled": solver.unscheduled,
        }
        _write_generation_report(db, {**payload, "status": "failed"})
        return (payload, 400)

    if progress_callback:
        progress_callback({
            "stage": "saving_timetable",
            "message": "Saving generated timetable and updating versioned slots",
            "percent": 99.5,
            "details": {
                "lecture_slots": len(solver.assignments),
                "lab_slots": len(solver.lab_assignments),
            },
        })
    try:
        version_id = solver.save_to_db()
        db.slot_claims.delete_many({})
    except (BulkWriteError, ValueError) as exc:
        payload = {
            "error": "Could not save timetable due to duplicate slot allocation",
            "detail": str(exc),
            "diagnostic": None,
            "diagnostics": [],
            "coverage_report": solver.coverage_report,
            "elapsed_seconds": elapsed_seconds,
            "partial": bool(solver.partial),
            "unscheduled": solver.unscheduled,
        }
        _write_generation_report(db, {**payload, "status": "failed"})
        return (payload, 400)

    iid = next_seq(db, "change_log")
    db.change_log.insert_one({
        "int_id":      iid,
        "change_type": "generate_partial" if solver.partial else "generate",
        "description": f"{len(solver.assignments)} lecture/tutorial slots + {len(solver.lab_assignments)} lab periods generated",
        "resolved_by": "solver",
        "success":     1,
        "changed_by":  changed_by,
        "created_at":  datetime.now().isoformat(),
        "affected":    "[]",
        "reason":      "",
    })
    diagnostics = _diagnostic_dicts(([strict_failure] if (not ok and allow_partial and strict_failure) else []) + list(solver.diagnostics or []))
    payload = {
        "ok": True,
        "version_id": version_id,
        "slots": len(solver.assignments),
        "lab_slots": len(solver.lab_assignments),
        "coverage_report": solver.coverage_report,
        "elapsed_seconds": elapsed_seconds,
        "partial": bool(solver.partial),
        "unscheduled": solver.unscheduled,
        "diagnostic": strict_failure.__dict__ if (not ok and allow_partial and strict_failure) else None,
        "diagnostics": diagnostics,
        "message": (
            f"Generated partial timetable with {len(solver.assignments)} lecture/tutorial slots and {len(solver.lab_assignments)} lab periods."
            if solver.partial
            else f"Generated {len(solver.assignments)} lecture/tutorial slots and {len(solver.lab_assignments)} lab periods."
        ),
    }
    _write_generation_report(db, {**payload, "status": "completed"})
    return (payload, 200)


def _run_generation_job(job_id, changed_by, allow_partial):
    started = time.perf_counter()

    def progress_callback(payload):
        _update_generation_job(
            job_id,
            status="running",
            stage=payload.get("stage"),
            message=payload.get("message", ""),
            percent=payload.get("percent", 0),
            details=payload.get("details") or {},
            elapsed_seconds=round(time.perf_counter() - started, 3),
        )

    try:
        with app.app_context():
            progress_callback({"stage": "starting", "message": "Loading current timetable data and rules", "percent": 1})
            payload, status_code = _execute_generation(changed_by=changed_by, progress_callback=progress_callback, allow_partial=allow_partial)
    except Exception as exc:
        _update_generation_job(
            job_id,
            status="failed",
            stage="failed",
            message=f"Generation crashed: {exc}",
            percent=100,
            elapsed_seconds=round(time.perf_counter() - started, 3),
            diagnostic=None,
            diagnostics=[],
            coverage_report={},
            result={"error": str(exc)},
        )
        return

    final_elapsed = payload.get("elapsed_seconds", round(time.perf_counter() - started, 3))
    if status_code == 200:
        _update_generation_job(
            job_id,
            status="completed",
            stage="completed",
            message=payload.get("message", "Generation completed"),
            percent=100,
            elapsed_seconds=final_elapsed,
            diagnostic=payload.get("diagnostic"),
            diagnostics=payload.get("diagnostics") or [],
            coverage_report=payload.get("coverage_report") or {},
            version_id=payload.get("version_id"),
            slots=payload.get("slots", 0),
            lab_slots=payload.get("lab_slots", 0),
            partial=payload.get("partial", False),
            unscheduled=payload.get("unscheduled") or [],
            result=payload,
        )
        return

    diagnostic = payload.get("diagnostic")
    _update_generation_job(
        job_id,
        status="failed",
        stage=(diagnostic or {}).get("phase") or "failed",
        message=payload.get("error", "Generation failed"),
        percent=100,
        elapsed_seconds=final_elapsed,
        diagnostic=diagnostic,
        diagnostics=payload.get("diagnostics") or [],
        coverage_report=payload.get("coverage_report") or {},
        partial=payload.get("partial", False),
        unscheduled=payload.get("unscheduled") or [],
        result=payload,
    )

# ─────────────────────────────────────────
#  Config helpers
# ─────────────────────────────────────────
def get_config(key, default=None):
    db  = get_db()
    row = db.config.find_one({"key": key})
    if row:
        try:    return json.loads(row['value'])
        except: return row['value']
    return default

def set_config(key, value):
    db = get_db()
    db.config.update_one(
        {"key": key},
        {"$set": {"value": json.dumps(value) if not isinstance(value, str) else value}},
        upsert=True
    )

# ─────────────────────────────────────────
#  Auth decorators
# ─────────────────────────────────────────
def login_required(f):
    @wraps(f)
    def dec(*a, **kw):
        if not _current_user():
            return jsonify({"error": "Not authenticated"}), 401
        return f(*a, **kw)
    return dec

def coordinator_only(f):
    @wraps(f)
    def dec(*a, **kw):
        if _current_user_value('role') != 'coordinator':
            return jsonify({"error": "Coordinator access required"}), 403
        return f(*a, **kw)
    return dec

@app.after_request
def add_headers(resp):
    resp.headers['X-Content-Type-Options'] = 'nosniff'
    return resp

# ─────────────────────────────────────────
#  Frontend serving
# ─────────────────────────────────────────
@app.route('/', defaults={'path': ''})
@app.route('/<path:path>')
def serve(path):
    if path.startswith('api/'):
        return jsonify({"error": "Not found"}), 404
    static = os.path.join(FRONTEND_DIR, 'static', path)
    if path and os.path.isfile(static):
        return send_from_directory(os.path.join(FRONTEND_DIR, 'static'), path)
    return send_from_directory(FRONTEND_DIR, 'index.html')

# ══════════════════════════════════════════
#  AUTH
# ══════════════════════════════════════════
@app.route('/api/auth/login', methods=['POST'])
def login():
    d  = request.get_json(silent=True) or {}
    em = (d.get('email') or '').strip().lower()
    pw = d.get('password') or ''
    if not em or not pw:
        return jsonify({"error": "Email and password required"}), 400
    db   = get_db()
    user = db.users.find_one({"email": em, "password": hash_pw(pw)})
    if not user:
        return jsonify({"error": "Invalid email or password"}), 401
    session.permanent = True
    session.update({
        "user_id":    str(user['int_id']),
        "role":       user['role'],
        "name":       user['name'],
        "teacher_id": user.get('teacher_id'),
        "email":      user.get('email'),
    })
    payload = _sanitize_user(user)
    payload["auth_token"] = _issue_auth_token(user)
    return jsonify(payload)

@app.route('/api/auth/logout', methods=['POST'])
def logout():
    _revoke_auth_token(_auth_token_from_request())
    session.clear()
    return jsonify({"ok": True})

@app.route('/api/auth/me')
def me():
    user = _current_user()
    if not user:
        return jsonify({"error": "Not authenticated"}), 401
    return jsonify(user)

@app.route('/api/auth/register', methods=['POST'])
def register():
    d    = request.get_json(silent=True) or {}
    role = d.get('role', 'faculty')
    db   = get_db()
    if role == 'coordinator':
        existing = db.users.count_documents({"role": "coordinator"})
        if existing and _current_user_value('role') != 'coordinator':
            return jsonify({"error": "Coordinator account already exists. Contact admin."}), 403
    em   = (d.get('email') or '').strip().lower()
    pw   = d.get('password') or ''
    name = (d.get('name') or '').strip()
    if not em or not pw or not name:
        return jsonify({"error": "Name, email and password required"}), 400
    if len(pw) < 6:
        return jsonify({"error": "Password must be at least 6 characters"}), 400
    if db.users.find_one({"email": em}):
        return jsonify({"error": "Email already registered"}), 409
    uid = next_seq(db, "users")
    db.users.insert_one({
        "int_id": uid, "email": em, "password": hash_pw(pw),
        "role": role, "name": name, "teacher_id": None,
    })
    return jsonify({"ok": True, "message": "Account created. You can now sign in."})

@app.route('/api/auth/change-password', methods=['POST'])
@login_required
def change_password():
    d      = request.get_json(silent=True) or {}
    old_pw = d.get('old_password', '')
    new_pw = d.get('new_password', '')
    if len(new_pw) < 6:
        return jsonify({"error": "Password must be at least 6 characters"}), 400
    db   = get_db()
    user = db.users.find_one({"int_id": int(_current_user_value('int_id')), "password": hash_pw(old_pw)})
    if not user:
        return jsonify({"error": "Current password is incorrect"}), 400
    db.users.update_one({"int_id": int(_current_user_value('int_id'))}, {"$set": {"password": hash_pw(new_pw)}})
    return jsonify({"ok": True, "message": "Password changed successfully"})

# ══════════════════════════════════════════
#  CONFIG
# ══════════════════════════════════════════
@app.route('/api/config', methods=['GET'])
@login_required
def get_all_config():
    db   = get_db()
    rows = db.config.find()
    result = {}
    for r in rows:
        try:    result[r['key']] = json.loads(r['value'])
        except: result[r['key']] = r['value']
    return jsonify(result)

@app.route('/api/config', methods=['PUT'])
@login_required
@coordinator_only
def update_config():
    d       = request.json or {}
    allowed = (
        'period_ids','period_labels','days','department_name','academic_year',
        'semester','break_after_periods','timetable_orientation'
    )
    for k, v in d.items():
        if k in allowed:
            set_config(k, v)
    return jsonify({"ok": True, "message": "Settings saved"})

# ══════════════════════════════════════════
#  TEACHERS
# ══════════════════════════════════════════
@app.route('/api/teachers', methods=['GET'])
@login_required
def get_teachers():
    db   = get_db()
    rows = list(db.teachers.find({}, {"_id": 0}).sort("name", 1))
    for t in rows:
        if isinstance(t.get('unavailable'), str):
            try:    t['unavailable'] = json.loads(t['unavailable'])
            except: t['unavailable'] = {}
    return jsonify(rows)

@app.route('/api/teachers', methods=['POST'])
@login_required
@coordinator_only
def add_teacher():
    d     = request.json or {}
    name  = (d.get('name') or '').strip()
    short = (d.get('short_name') or '').strip()
    email = (d.get('email') or '').strip().lower()
    if not name or not short:
        return jsonify({"error": "Name and short name are required"}), 400
    tid = (d.get('id') or short.upper().replace(' ', ''))[:20]
    db  = get_db()
    if db.teachers.find_one({"id": tid}):
        return jsonify({"error": f"Teacher ID '{tid}' already exists"}), 409
    if email and db.teachers.find_one({"email": email}):
        return jsonify({"error": "Email already in use"}), 409
    max_day = max(1, int(d.get('max_hrs_per_day', 4)))
    max_week = max(1, int(d.get('max_hrs_per_week', 18)))
    db.teachers.insert_one({
        "id": tid, "name": name, "short_name": short, "email": email,
        "max_hrs_per_day":  max_day,
        "max_hrs_per_week": max_week,
        "unavailable":      d.get('unavailable', {}),
    })
    if email:
        uid = next_seq(db, "users")
        db.users.update_one(
            {"email": email},
            {"$setOnInsert": {
                "int_id": uid, "email": email, "password": hash_pw("faculty123"),
                "role": "faculty", "name": name, "teacher_id": tid,
            }},
            upsert=True
        )
    return jsonify({"id": tid, "message": f"Teacher added. Login: {email} / faculty123"})

@app.route('/api/teachers/<tid>', methods=['PUT'])
@login_required
@coordinator_only
def update_teacher(tid):
    d  = request.json or {}
    db = get_db()
    max_day = max(1, int(d.get('max_hrs_per_day', 4)))
    max_week = max(1, int(d.get('max_hrs_per_week', 18)))
    db.teachers.update_one({"id": tid}, {"$set": {
        "name":            d.get('name'),
        "short_name":      d.get('short_name'),
        "email":           d.get('email', ''),
        "max_hrs_per_day":  max_day,
        "max_hrs_per_week": max_week,
        "unavailable":      d.get('unavailable', {}),
    }})
    return jsonify({"ok": True})

@app.route('/api/teachers/<tid>', methods=['DELETE'])
@login_required
@coordinator_only
def delete_teacher(tid):
    db   = get_db()
    used = db.timetable_slots.count_documents(active_timetable_filter(db, {"teacher_id": tid}))
    if used:
        return jsonify({"error": f"Cannot delete: teacher has {used} active timetable slots"}), 400
    db.batch_teachers.delete_many({"teacher_id": tid})
    db.teachers.delete_one({"id": tid})
    db.users.delete_many({"teacher_id": tid})
    return jsonify({"ok": True})

# ══════════════════════════════════════════
#  ROOMS
# ══════════════════════════════════════════
@app.route('/api/rooms', methods=['GET'])
@login_required
def get_rooms():
    db   = get_db()
    rows = list(db.rooms.find({}, {"_id": 0}).sort([("room_type", 1), ("name", 1)]))
    return jsonify(rows)

@app.route('/api/rooms', methods=['POST'])
@login_required
@coordinator_only
def add_room():
    d         = request.json or {}
    name      = (d.get('name') or '').strip()
    room_type = d.get('room_type', 'lecture')
    if not name:
        return jsonify({"error": "Room name is required"}), 400
    if room_type not in ('lecture', 'lab'):
        return jsonify({"error": "room_type must be lecture or lab"}), 400
    rid = (d.get('id') or name.upper().replace(' ', ''))[:20]
    db  = get_db()
    if db.rooms.find_one({"id": rid}):
        return jsonify({"error": f"Room ID '{rid}' already exists"}), 409
    db.rooms.insert_one({"id": rid, "name": name, "room_type": room_type,
                          "capacity": int(d.get('capacity', 60))})
    return jsonify({"id": rid})

@app.route('/api/rooms/<rid>', methods=['DELETE'])
@login_required
@coordinator_only
def delete_room(rid):
    db   = get_db()
    used = db.timetable_slots.count_documents(active_timetable_filter(db, {"room_id": rid}))
    if used:
        return jsonify({"error": f"Cannot delete: room has {used} active slots"}), 400
    db.rooms.delete_one({"id": rid})
    return jsonify({"ok": True})

# ══════════════════════════════════════════
#  SUBJECTS
# ══════════════════════════════════════════
@app.route('/api/subjects', methods=['GET'])
@login_required
def get_subjects():
    db   = get_db()
    rows = list(db.subjects.find({}, {"_id": 0}).sort("name", 1))
    return jsonify(rows)

@app.route('/api/subjects', methods=['POST'])
@login_required
@coordinator_only
def add_subject():
    d     = request.json or {}
    name  = (d.get('name') or '').strip()
    short = (d.get('short_name') or '').strip()
    if not name or not short:
        return jsonify({"error": "Name and short name required"}), 400
    has_lab = bool(d.get('has_lab'))
    lab_hrs = int(d.get('lab_hours_per_week', 0))
    if has_lab and lab_hrs > 0 and lab_hrs % 2 != 0:
        return jsonify({"error": "Lab hours per week must be even (each session = 2 hrs)"}), 400
    if has_lab and not d.get('lab_teacher_id'):
        return jsonify({"error": "Default lab teacher required when lab is enabled"}), 400
    sid = (d.get('id') or short.upper().replace(' ', ''))[:20]
    db  = get_db()
    if db.subjects.find_one({"id": sid}):
        return jsonify({"error": f"Subject ID '{sid}' already exists"}), 409
    db.subjects.insert_one({
        "id": sid, "name": name, "short_name": short,
        "code":               d.get('code', ''),
        "teacher_id":         d.get('teacher_id'),
        "lectures_per_week":  int(d.get('lectures_per_week', 3)),
        "has_lab":            1 if has_lab else 0,
        "lab_hours_per_week": lab_hrs,
        "lab_teacher_id":     d.get('lab_teacher_id') if has_lab else None,
        "lab_room_id":        d.get('lab_room_id') if has_lab else None,
        "has_tutorial":       1 if d.get('has_tutorial') else 0,
        "tutorials_per_week": int(d.get('tutorials_per_week', 0)),
    })
    return jsonify({"id": sid})

@app.route('/api/subjects/<sid>', methods=['PUT'])
@login_required
@coordinator_only
def update_subject(sid):
    d       = request.json or {}
    has_lab = bool(d.get('has_lab'))
    db      = get_db()
    db.subjects.update_one({"id": sid}, {"$set": {
        "name": d.get('name'), "short_name": d.get('short_name'),
        "code": d.get('code', ''), "teacher_id": d.get('teacher_id'),
        "lectures_per_week":  int(d.get('lectures_per_week', 3)),
        "has_lab":            1 if has_lab else 0,
        "lab_hours_per_week": int(d.get('lab_hours_per_week', 0)),
        "lab_teacher_id":     d.get('lab_teacher_id') if has_lab else None,
        "lab_room_id":        d.get('lab_room_id') if has_lab else None,
        "has_tutorial":       1 if d.get('has_tutorial') else 0,
        "tutorials_per_week": int(d.get('tutorials_per_week', 0)),
    }})
    return jsonify({"ok": True})

@app.route('/api/subjects/<sid>', methods=['DELETE'])
@login_required
@coordinator_only
def delete_subject(sid):
    db   = get_db()
    used = db.timetable_slots.count_documents(active_timetable_filter(db, {"subject_id": sid}))
    if used:
        return jsonify({"error": f"Cannot delete: subject has {used} active slots"}), 400
    db.batch_teachers.delete_many({"subject_id": sid})
    db.division_subjects.delete_many({"subject_id": sid})
    db.subjects.delete_one({"id": sid})
    return jsonify({"ok": True})

# ══════════════════════════════════════════
#  DIVISIONS  (with named batches)
# ══════════════════════════════════════════
def _division_with_batches(div):
    """Attach subjects list and batches to a division dict."""
    db  = get_db()
    did = div['id']
    div['subjects'] = [
        r['subject_id']
        for r in db.division_subjects.find({"division_id": did}, {"_id": 0})
    ]
    div['batches'] = list(
        db.batches.find({"division_id": did}, {"_id": 0}).sort("id", 1)
    )
    return div

@app.route('/api/divisions', methods=['GET'])
@login_required
def get_divisions():
    db   = get_db()
    divs = list(db.divisions.find({}, {"_id": 0}).sort("id", 1))
    return jsonify([_division_with_batches(d) for d in divs])

@app.route('/api/divisions', methods=['POST'])
@login_required
@coordinator_only
def add_division():
    d    = request.json or {}
    name = (d.get('name') or '').strip()
    if not name:
        return jsonify({"error": "Division name is required"}), 400

    did         = (d.get('id') or name.upper().replace(' ', '').replace('DIVISION', ''))[:10]
    size        = int(d.get('size', 60))
    num_batches = max(1, min(10, int(d.get('num_batches', 4))))

    # batch_names: list of custom names; pad/trim to match num_batches
    batch_names = d.get('batch_names') or []
    if not isinstance(batch_names, list):
        batch_names = []
    # Fill missing names with defaults
    for i in range(len(batch_names), num_batches):
        batch_names.append(f"Batch {i + 1}")
    batch_names = [str(n).strip() or f"Batch {i+1}" for i, n in enumerate(batch_names[:num_batches])]

    db = get_db()
    if db.divisions.find_one({"id": did}):
        return jsonify({"error": f"Division '{did}' already exists"}), 409

    db.divisions.insert_one({
        "id": did, "name": name, "room_id": d.get('room_id') or None, "size": size
    })
    db.division_subjects.delete_many({"division_id": did})
    for sid in d.get('subjects', []):
        db.division_subjects.update_one(
            {"division_id": did, "subject_id": sid},
            {"$setOnInsert": {"division_id": did, "subject_id": sid}},
            upsert=True
        )
    db.batches.delete_many({"division_id": did})
    batch_size = max(1, size // num_batches)
    for i, bname in enumerate(batch_names, 1):
        bid = f"{did}{i}"
        db.batches.insert_one({
            "id":          bid,
            "name":        bname,
            "division_id": did,
            "size":        batch_size,
        })
    return jsonify({"id": did})

@app.route('/api/divisions/<did>', methods=['PUT'])
@login_required
@coordinator_only
def update_division(did):
    d  = request.json or {}
    db = get_db()
    db.divisions.update_one({"id": did}, {"$set": {
        "name":    d.get('name'),
        "room_id": d.get('room_id') or None,
        "size":    int(d.get('size', 60)),
    }})
    db.division_subjects.delete_many({"division_id": did})
    for sid in d.get('subjects', []):
        db.division_subjects.update_one(
            {"division_id": did, "subject_id": sid},
            {"$setOnInsert": {"division_id": did, "subject_id": sid}},
            upsert=True
        )
    return jsonify({"ok": True})

@app.route('/api/divisions/<did>', methods=['DELETE'])
@login_required
@coordinator_only
def delete_division(did):
    db = get_db()
    batch_ids = [b['id'] for b in db.batches.find({"division_id": did}, {"id": 1})]
    slot_ids  = [s['int_id'] for s in db.timetable_slots.find(active_timetable_filter(db, {"division_id": did}), {"int_id": 1})]
    if slot_ids:
        db.slot_claims.delete_many({"slot_id": {"$in": slot_ids}})
    db.timetable_slots.delete_many(active_timetable_filter(db, {"division_id": did}))
    if batch_ids:
        db.batch_teachers.delete_many({"batch_id": {"$in": batch_ids}})
    db.batches.delete_many({"division_id": did})
    db.division_subjects.delete_many({"division_id": did})
    db.divisions.delete_one({"id": did})
    return jsonify({"ok": True})

# ══════════════════════════════════════════
#  BATCHES  (management endpoints)
# ══════════════════════════════════════════
@app.route('/api/batches', methods=['GET'])
@login_required
def get_batches():
    """List all batches (optionally filtered by division_id)."""
    db  = get_db()
    flt = {}
    div = request.args.get('division')
    if div:
        flt['division_id'] = div
    batches = list(db.batches.find(flt, {"_id": 0}).sort([("division_id", 1), ("id", 1)]))
    return jsonify(batches)

@app.route('/api/batches/<bid>', methods=['PUT'])
@login_required
@coordinator_only
def update_batch(bid):
    """Rename a batch or change its size."""
    d  = request.json or {}
    db = get_db()
    upd = {}
    if d.get('name'):
        upd['name'] = d['name'].strip()
    if d.get('size'):
        upd['size'] = int(d['size'])
    if not upd:
        return jsonify({"error": "Nothing to update"}), 400
    db.batches.update_one({"id": bid}, {"$set": upd})
    return jsonify({"ok": True})

@app.route('/api/batches/<bid>', methods=['DELETE'])
@login_required
@coordinator_only
def delete_batch(bid):
    db   = get_db()
    used = db.timetable_slots.count_documents(active_timetable_filter(db, {"batch_id": bid}))
    if used:
        return jsonify({"error": f"Cannot delete: batch has {used} active timetable slots"}), 400
    db.batch_teachers.delete_many({"batch_id": bid})
    db.batches.delete_one({"id": bid})
    return jsonify({"ok": True})

@app.route('/api/divisions/<did>/batches', methods=['POST'])
@login_required
@coordinator_only
def add_batch_to_division(did):
    """Add a new named batch to an existing division."""
    d  = request.json or {}
    db = get_db()
    div = db.divisions.find_one({"id": did})
    if not div:
        return jsonify({"error": "Division not found"}), 404
    existing = db.batches.count_documents({"division_id": did})
    bid  = f"{did}{existing + 1}"
    name = (d.get('name') or f"Batch {existing + 1}").strip()
    size = int(d.get('size') or div.get('size', 60))
    db.batches.insert_one({"id": bid, "name": name, "division_id": did, "size": size})
    return jsonify({"id": bid, "name": name})

# ══════════════════════════════════════════
#  BATCH TEACHERS
# ══════════════════════════════════════════
@app.route('/api/batch-teachers', methods=['GET'])
@login_required
def get_batch_teachers():
    db  = get_db()
    div = request.args.get('division')
    if div:
        batch_ids = [b['id'] for b in db.batches.find({"division_id": div}, {"id": 1})]
        rows = list(db.batch_teachers.find({"batch_id": {"$in": batch_ids}}, {"_id": 0}))
    else:
        rows = list(db.batch_teachers.find({}, {"_id": 0}))
    # Attach division_id from batches
    batch_map = {b['id']: b['division_id'] for b in db.batches.find({}, {"id": 1, "division_id": 1})}
    batch_name_map = {b['id']: b.get('name', b['id']) for b in db.batches.find({}, {"id": 1, "name": 1})}
    for r in rows:
        r['division_id']  = batch_map.get(r.get('batch_id'), '')
        r['batch_name']   = batch_name_map.get(r.get('batch_id'), r.get('batch_id', ''))
    return jsonify(rows)

@app.route('/api/batch-teachers', methods=['POST'])
@login_required
@coordinator_only
def set_batch_teacher():
    d          = request.json or {}
    batch_id   = d.get('batch_id')
    subject_id = d.get('subject_id')
    teacher_id = d.get('teacher_id')
    if not all([batch_id, subject_id, teacher_id]):
        return jsonify({"error": "batch_id, subject_id, teacher_id all required"}), 400
    db = get_db()
    iid = next_seq(db, "batch_teachers")
    db.batch_teachers.update_one(
        {"batch_id": batch_id, "subject_id": subject_id},
        {"$set": {"teacher_id": teacher_id, "int_id": iid}},
        upsert=True
    )
    return jsonify({"ok": True})

@app.route('/api/batch-teachers/<int:btid>', methods=['DELETE'])
@login_required
@coordinator_only
def delete_batch_teacher(btid):
    db = get_db()
    db.batch_teachers.delete_one({"int_id": btid})
    return jsonify({"ok": True})

# ══════════════════════════════════════════
#  GENERATE
# ══════════════════════════════════════════
@app.route('/api/generate', methods=['POST'])
@login_required
@coordinator_only
def generate():
    active = _get_active_generation_job()
    if active:
        return jsonify({
            "error": "Timetable generation is already running",
            "job": active,
        }), 409
    body = request.get_json(silent=True) or {}
    allow_partial = bool(body.get('allow_partial', True))
    payload, status_code = _execute_generation(changed_by=_current_user_value('name'), allow_partial=allow_partial)
    return jsonify(payload), status_code


@app.route('/api/generate/start', methods=['POST'])
@login_required
@coordinator_only
def start_generate_job():
    active = _get_active_generation_job()
    if active:
        return jsonify({"ok": True, "reused": True, "job": active}), 202
    allow_partial = bool((request.get_json(silent=True) or {}).get("allow_partial", True))
    job = _create_generation_job(_current_user_value('name'))
    _update_generation_job(job["job_id"], allow_partial=allow_partial)
    worker = threading.Thread(target=_run_generation_job, args=(job["job_id"], _current_user_value('name'), allow_partial), daemon=True)
    worker.start()
    return jsonify({"ok": True, "job": _get_generation_job(job["job_id"])}), 202


@app.route('/api/generate/active')
@login_required
@coordinator_only
def active_generate_job():
    return jsonify({"job": _get_active_generation_job()})


@app.route('/api/generate/jobs/<job_id>')
@login_required
@coordinator_only
def get_generate_job(job_id):
    job = _get_generation_job(job_id)
    if not job:
        return jsonify({"error": "Generation job not found"}), 404
    return jsonify(job)

# ══════════════════════════════════════════
#  TIMETABLE
# ══════════════════════════════════════════
@app.route('/api/timetable')
@login_required
def timetable():
    div = request.args.get('division')
    return jsonify(get_timetable(div))

@app.route('/api/timetable/meta')
@login_required
def timetable_meta():
    db       = get_db()
    teachers = {r['id']: r for r in db.teachers.find({}, {"_id": 0})}
    rooms    = {r['id']: r for r in db.rooms.find({}, {"_id": 0})}
    subjects = {r['id']: r for r in db.subjects.find({}, {"_id": 0})}
    divs_raw = list(db.divisions.find({}, {"_id": 0}).sort("id", 1))
    divisions = {}
    for div in divs_raw:
        did = div['id']
        div['subjects'] = [r['subject_id'] for r in db.division_subjects.find({"division_id": did})]
        # Fetch batches for this division
        div['batches'] = [r for r in db.batches.find({"division_id": did}, {"_id": 0}).sort("id", 1)]
        divisions[did] = div
    return jsonify({
        "teachers":      teachers,
        "rooms":         rooms,
        "subjects":      subjects,
        "divisions":     divisions,
        "period_ids":    get_config('period_ids', ["P1","P2","P3","P4","P5","P6"]),
        "period_labels": get_config('period_labels', {}),
        "days":          get_config('days', ["Monday","Tuesday","Wednesday","Thursday","Friday"]),
        "break_after_periods": get_config('break_after_periods', ["P2","P4"]),
        "timetable_orientation": get_config('timetable_orientation', 'days_horizontal'),
        "department":    get_config('department_name', ''),
        "academic_year": get_config('academic_year', ''),
        "semester":      get_config('semester', ''),
        "startup_report": get_startup_report(),
    })

@app.route('/api/timetable/coverage')
@login_required
def timetable_coverage():
    db = get_db()
    solver = TimetableSolver(period_ids=get_config('period_ids', ["P1","P2","P3","P4","P5","P6"]),
                             days=get_config('days', ["Monday","Tuesday","Wednesday","Thursday","Friday"]),
                             db=db)
    solver.state = build_state_from_active_timetable(db, solver.context)
    _, report = solver.validate_required_coverage()
    return jsonify(report)

@app.route('/api/timetable/diagnostics')
@login_required
def timetable_diagnostics():
    return jsonify([d.__dict__ for d in validate_active_timetable(get_db())])

@app.route('/api/timetable/problems')
@login_required
def timetable_problems():
    return jsonify(_build_timetable_problem_report(get_db()))

@app.route('/api/timetable/clear', methods=['POST'])
@login_required
@coordinator_only
def clear_timetable():
    db = get_db()
    db.slot_claims.delete_many({})
    clear_active_timetable(db)
    return jsonify({"ok": True})

# ══════════════════════════════════════════
#  SLOTS
# ══════════════════════════════════════════
@app.route('/api/slots')
@login_required
def get_slots():
    db  = get_db()
    flt = {"slot_type": {"$ne": "lab"}}
    div = request.args.get('division')
    if div:
        flt['division_id'] = div
    rows = list(db.timetable_slots.find(active_timetable_filter(db, flt), {"_id": 0}).sort([("division_id",1),("day",1),("period_id",1)]))
    return jsonify(rows)

@app.route('/api/slots/<int:slot_id>/move', methods=['POST'])
@login_required
@coordinator_only
def api_move(slot_id):
    d = request.json or {}
    if not d.get('day') or not d.get('period_id'):
        return jsonify({"error": "day and period_id required"}), 400
    ok, msg = move_slot(slot_id, d['day'], d['period_id'], changed_by=_current_user_value('name'))
    return jsonify({"ok": ok, "message": msg}), (200 if ok else 400)

@app.route('/api/slots/swap', methods=['POST'])
@login_required
@coordinator_only
def api_swap():
    d = request.json or {}
    if not d.get('slot_id_1') or not d.get('slot_id_2'):
        return jsonify({"error": "slot_id_1 and slot_id_2 required"}), 400
    ok, msg = swap_slots(int(d['slot_id_1']), int(d['slot_id_2']), changed_by=_current_user_value('name'))
    return jsonify({"ok": ok, "message": msg}), (200 if ok else 400)

@app.route('/api/slots/validate-swap', methods=['POST'])
@login_required
@coordinator_only
def api_validate_swap():
    d = request.json or {}
    if not d.get('slot_id_1') or not d.get('slot_id_2'):
        return jsonify({"error": "slot_id_1 and slot_id_2 required"}), 400
    ok, msg, payload = validate_or_apply_swap(int(d['slot_id_1']), int(d['slot_id_2']), apply_change=False, changed_by=_current_user_value('name'))
    return jsonify({"ok": ok, "message": msg, **payload}), (200 if ok else 400)

@app.route('/api/slots/<int:slot_id>/lock', methods=['POST'])
@login_required
@coordinator_only
def api_lock(slot_id):
    locked = (request.json or {}).get('locked', True)
    ok, msg = lock_slot(slot_id, locked)
    return jsonify({"ok": ok, "message": msg})

@app.route('/api/slots/<int:slot_id>/teacher', methods=['PUT'])
@login_required
@coordinator_only
def api_change_teacher(slot_id):
    d = request.json or {}
    if not d.get('teacher_id'):
        return jsonify({"error": "teacher_id required"}), 400
    ok, msg = change_teacher_for_slot(slot_id, d['teacher_id'], changed_by=_current_user_value('name'))
    return jsonify({"ok": ok, "message": msg}), (200 if ok else 400)


@app.route('/api/slots/validate-cell-move', methods=['POST'])
@login_required
@coordinator_only
def api_validate_cell_move():
    d = request.json or {}
    slot_ids = [int(slot_id) for slot_id in (d.get('slot_ids') or []) if str(slot_id).strip()]
    if not slot_ids or not d.get('day') or not d.get('period_id'):
        return jsonify({"error": "slot_ids, day, and period_id are required"}), 400
    ok, msg, payload = validate_or_apply_cell_move(slot_ids, d['day'], d['period_id'], apply_change=False, changed_by=_current_user_value('name'))
    return jsonify({"ok": ok, "message": msg, **payload}), (200 if ok else 400)


@app.route('/api/slots/cell-move', methods=['POST'])
@login_required
@coordinator_only
def api_cell_move():
    d = request.json or {}
    slot_ids = [int(slot_id) for slot_id in (d.get('slot_ids') or []) if str(slot_id).strip()]
    if not slot_ids or not d.get('day') or not d.get('period_id'):
        return jsonify({"error": "slot_ids, day, and period_id are required"}), 400
    ok, msg, payload = validate_or_apply_cell_move(slot_ids, d['day'], d['period_id'], apply_change=True, changed_by=_current_user_value('name'))
    return jsonify({"ok": ok, "message": msg, **payload}), (200 if ok else 400)

# ══════════════════════════════════════════
#  ABSENCE
# ══════════════════════════════════════════
@app.route('/api/teacher-absent', methods=['POST'])
@login_required
def api_absent():
    d   = request.json or {}
    tid = d.get('teacher_id') or _current_user_value('teacher_id')
    if not tid:
        return jsonify({"error": "teacher_id required"}), 400
    if not d.get('day'):
        return jsonify({"error": "day required"}), 400
    if _current_user_value('role') == 'faculty' and tid != _current_user_value('teacher_id'):
        return jsonify({"error": "You can only report your own absence"}), 403
    result = teacher_absent(tid, d['day'], d.get('period_id'), changed_by=_current_user_value('name'))
    return jsonify(result), (200 if result.get('success') else 400)

# ══════════════════════════════════════════
#  CONSTRAINTS
# ══════════════════════════════════════════
@app.route('/api/constraints', methods=['GET'])
@login_required
def get_constraints():
    db   = get_db()
    rows = list(db.user_constraints.find({}, {"_id": 0}).sort("created_at", -1))
    return jsonify(rows)

@app.route('/api/constraints/meta')
@login_required
def constraint_meta():
    return jsonify(constraints_ui_meta(
        get_config('period_ids', ["P1","P2","P3","P4","P5","P6"]),
        get_config('days', ["Monday","Tuesday","Wednesday","Thursday","Friday"]),
    ))

@app.route('/api/constraints/validate', methods=['POST'])
@login_required
def validate_constraint():
    payload = request.json or {}
    normalized, errors = normalize_constraint_document(payload, strict=False)
    return jsonify({
        "ok": normalized is not None and not errors,
        "errors": errors,
        "normalized": normalized,
    }), (200 if normalized is not None else 400)

@app.route('/api/constraints', methods=['POST'])
@login_required
@coordinator_only
def post_constraint():
    d = request.json or {}
    if not d.get('constraint_type'):
        return jsonify({"error": "constraint_type required"}), 400
    ok, msg = add_constraint(d, changed_by=_current_user_value('name'))
    return jsonify({"ok": ok, "message": msg}), (200 if ok else 400)

@app.route('/api/constraints/<int:cid>', methods=['PUT'])
@login_required
@coordinator_only
def put_constraint(cid):
    ok, msg = update_constraint(cid, request.json or {}, changed_by=_current_user_value('name'))
    return jsonify({"ok": ok, "message": msg}), (200 if ok else 400)

@app.route('/api/constraints/<int:cid>', methods=['DELETE'])
@login_required
@coordinator_only
def delete_constraint(cid):
    db = get_db()
    db.user_constraints.delete_one({"int_id": cid})
    return jsonify({"ok": True})

# ══════════════════════════════════════════
#  NOTIFICATIONS
# ══════════════════════════════════════════
@app.route('/api/notifications')
@login_required
def notifications():
    unread = request.args.get('unread') == 'true'
    notifs = get_notifications(teacher_id=_current_user_value('teacher_id'),
                                role=_current_user_value('role'), unread_only=unread)
    for n in notifs:
        if isinstance(n.get('data'), str):
            try:    n['data'] = json.loads(n['data'])
            except: n['data'] = {}
    return jsonify(notifs)

@app.route('/api/notifications/unread-count')
@login_required
def notif_count():
    notifs = get_notifications(teacher_id=_current_user_value('teacher_id'),
                                role=_current_user_value('role'), unread_only=True)
    return jsonify({"count": len(notifs)})

@app.route('/api/notifications/<int:nid>/read', methods=['POST'])
@login_required
def mark_read(nid):
    db = get_db()
    db.notifications.update_one({"$and": [{"int_id": nid}, _notification_visibility_filter()]}, {"$set": {"is_read": 1}})
    return jsonify({"ok": True})

@app.route('/api/notifications/<int:nid>/claim', methods=['POST'])
@login_required
def api_claim(nid):
    if _current_user_value('role') != 'faculty':
        return jsonify({"error": "Only faculty can claim slots"}), 403
    tid = _current_user_value('teacher_id')
    if not tid:
        return jsonify({"error": "Your account is not linked to a teacher profile"}), 400
    ok, msg = claim_slot(nid, tid, changed_by=_current_user_value('name'), assigned_by_role='faculty')
    return jsonify({"ok": ok, "message": msg}), (200 if ok else 400)


@app.route('/api/notifications/<int:nid>/assign', methods=['POST'])
@login_required
@coordinator_only
def api_assign_notification(nid):
    d = request.get_json(silent=True) or {}
    teacher_id = (d.get('teacher_id') or '').strip()
    if not teacher_id:
        return jsonify({"error": "teacher_id required"}), 400
    ok, msg = claim_slot(nid, teacher_id, changed_by=_current_user_value('name'), assigned_by_role='coordinator')
    return jsonify({"ok": ok, "message": msg}), (200 if ok else 400)

@app.route('/api/notifications/mark-all-read', methods=['POST'])
@login_required
def mark_all_read():
    db = get_db()
    db.notifications.update_many(_notification_visibility_filter(), {"$set": {"is_read": 1}})
    return jsonify({"ok": True})

# ══════════════════════════════════════════
#  FACULTY
# ══════════════════════════════════════════
@app.route('/api/my-schedule')
@login_required
def my_schedule():
    tid = _current_user_value('teacher_id')
    if not tid:
        return jsonify([])
    db   = get_db()
    rows = list(db.timetable_slots.find(active_timetable_filter(db, {"teacher_id": tid}), {"_id": 0}).sort([("day",1),("period_id",1)]))
    return jsonify(rows)

@app.route('/api/change-log')
@login_required
def changelog():
    return jsonify(get_change_log()[:100])

# ══════════════════════════════════════════
#  USERS
# ══════════════════════════════════════════
@app.route('/api/users')
@login_required
@coordinator_only
def get_users():
    db   = get_db()
    rows = list(db.users.find({}, {"_id": 0, "password": 0}).sort([("role",1),("name",1)]))
    for r in rows:
        r['id'] = r.get('int_id')
    return jsonify(rows)

@app.route('/api/users/<int:uid>/reset-password', methods=['POST'])
@login_required
@coordinator_only
def reset_password(uid):
    new_pw = (request.json or {}).get('password', 'faculty123')
    if len(new_pw) < 6:
        return jsonify({"error": "Password too short"}), 400
    db = get_db()
    db.users.update_one({"int_id": uid}, {"$set": {"password": hash_pw(new_pw)}})
    return jsonify({"ok": True, "message": f"Password reset to: {new_pw}"})

# ══════════════════════════════════════════
#  EXCEL EXPORT
# ══════════════════════════════════════════
@app.route('/api/export/excel')
@login_required
def export_excel():
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return jsonify({"error": "openpyxl not installed. Run: pip install openpyxl"}), 500

    db        = get_db()
    teachers  = {r['id']: r for r in db.teachers.find({}, {"_id": 0})}
    rooms     = {r['id']: r for r in db.rooms.find({}, {"_id": 0})}
    subjects  = {r['id']: r for r in db.subjects.find({}, {"_id": 0})}
    divisions = {r['id']: r for r in db.divisions.find({}, {"_id": 0}).sort("id",1)}
    all_slots = list(db.timetable_slots.find(active_timetable_filter(db), {"_id": 0}).sort([("division_id",1),("day",1),("period_id",1)]))
    change_logs = list(db.change_log.find({}, {"_id": 0}).sort("created_at", -1).limit(200))
    # batch name map
    batch_name_map = {b['id']: b.get('name', b['id']) for b in db.batches.find({}, {"id":1,"name":1})}

    period_ids    = get_config('period_ids', ["P1","P2","P3","P4","P5","P6"])
    period_labels = get_config('period_labels', {})
    days          = get_config('days', ["Monday","Tuesday","Wednesday","Thursday","Friday"])
    dept          = get_config('department_name', '')
    ay            = get_config('academic_year', '')
    sem           = get_config('semester', '')

    slot_map = {}
    for s in all_slots:
        key = (s['division_id'], s['day'], s['period_id'])
        slot_map.setdefault(key, []).append(s)

    BLUE  = "1E3A5F"; LBLUE = "2E6DAD"; WHITE = "FFFFFF"; LGRAY = "F4F6F9"
    def fill(c): return PatternFill("solid", fgColor=c)
    def thin(c="C5CDD8"):
        s = Side(style="thin", color=c)
        return Border(left=s, right=s, top=s, bottom=s)

    SCOL = ["DBEAFE","D1FAE5","FEF3C7","EDE9FE","CCFBF1","FFE4E6","FEF9C3","F3E8FF"]
    SACC = ["3B82F6","10B981","F59E0B","8B5CF6","14B8A6","F43F5E","EAB308","A855F7"]
    sci  = {}; scc = [0]
    def scol(sid):
        if sid not in sci: sci[sid] = scc[0] % len(SCOL); scc[0] += 1
        return SCOL[sci[sid]], SACC[sci[sid]]

    def lecture_lines(rows):
        lines = []
        for idx, slot in enumerate(sorted(rows, key=lambda row: (row.get("subject_id") or "", row.get("teacher_id") or "", row.get("room_id") or ""))):
            subj = subjects.get(slot["subject_id"], {"short_name": slot["subject_id"]})
            tchr = teachers.get(slot.get("teacher_id"), {"short_name": slot.get("teacher_id") or "OPEN"})
            room = rooms.get(slot.get("room_id"), {"name": slot.get("room_id") or "TBA"})
            tag = "TUT" if slot["slot_type"] == "tutorial" else "LEC"
            lines.extend([f"{subj['short_name']} [{tag}]", tchr["short_name"], room["name"]])
            if idx != len(rows) - 1:
                lines.append("")
        return "\n".join(lines)

    def lab_lines(rows):
        ordered = sorted(rows, key=lambda row: (batch_name_map.get(row.get("batch_id"), row.get("batch_id") or ""), row.get("room_id") or "", row.get("teacher_id") or ""))
        subject_names = sorted({(subjects.get(row["subject_id"], {"short_name": row["subject_id"]})["short_name"]) for row in ordered})
        lines = [f"[LAB] {' / '.join(subject_names)}"]
        seen_batches = set()
        for slot in ordered:
            batch_id = slot.get("batch_id")
            if batch_id in seen_batches:
                continue
            seen_batches.add(batch_id)
            subj = subjects.get(slot["subject_id"], {"short_name": slot["subject_id"]})
            tchr = teachers.get(slot.get("teacher_id"), {"short_name": slot.get("teacher_id") or "OPEN"})
            room = rooms.get(slot.get("room_id"), {"name": slot.get("room_id") or "TBA"})
            batch_name = batch_name_map.get(batch_id, batch_id or "Batch")
            lines.append(f"{batch_name}: {subj['short_name']} - {tchr['short_name']} @ {room['name']}")
        return "\n".join(lines)

    wb = Workbook()
    wb.remove(wb.active)

    for did, div in divisions.items():
        ws = wb.create_sheet(title=f"Div {did}")
        last = get_column_letter(len(period_ids)+1)
        ws.merge_cells(f"A1:{last}1")
        ws['A1'].value = f"Division {did} — {div['name']}  |  {dept}  |  {ay} {sem}"
        ws['A1'].font  = Font(name="Arial", size=12, bold=True, color=WHITE)
        ws['A1'].fill  = fill(BLUE)
        ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 20

        ws.cell(2,1).value = "Day"
        ws.cell(2,1).font=Font(name="Arial",size=9,bold=True,color=WHITE)
        ws.cell(2,1).fill=fill(BLUE)
        ws.cell(2,1).alignment=Alignment(horizontal="center",vertical="center")
        ws.cell(2,1).border=thin()
        ws.column_dimensions["A"].width = 12
        for ci, pid in enumerate(period_ids, 2):
            c=ws.cell(2,ci); c.value=period_labels.get(pid,pid)
            c.font=Font(name="Arial",size=9,bold=True,color=WHITE); c.fill=fill(LBLUE)
            c.alignment=Alignment(horizontal="center",vertical="center"); c.border=thin()
            ws.column_dimensions[get_column_letter(ci)].width=18
        ws.row_dimensions[2].height = 16

        for ri, day in enumerate(days):
            row = ri + 3; ws.row_dimensions[row].height = 74
            dc = ws.cell(row,1); dc.value=day
            dc.font=Font(name="Arial",size=9,bold=True,color=WHITE)
            dc.fill=fill(LBLUE); dc.alignment=Alignment(horizontal="center",vertical="center"); dc.border=thin()
            for ci, pid in enumerate(period_ids, 2):
                c     = ws.cell(row, ci); c.alignment=Alignment(vertical="top",wrap_text=True)
                cells = slot_map.get((did,day,pid),[])
                main  = [s for s in cells if s['slot_type']!='lab']
                labs  = [s for s in cells if s['slot_type']=='lab']
                if main:
                    bg,ac=scol(main[0]['subject_id'])
                    c.value=lecture_lines(main)
                    c.font=Font(name="Arial",size=9); c.fill=fill(bg)
                    c.border=Border(left=Side(style="medium",color=ac),right=Side(style="thin",color="C5CDD8"),
                                    top=Side(style="thin",color="C5CDD8"),bottom=Side(style="thin",color="C5CDD8"))
                elif labs:
                    c.value=lab_lines(labs)
                    c.font=Font(name="Arial",size=9,color="065F46"); c.fill=fill("D1FAE5")
                    c.border=Border(left=Side(style="medium",color="10B981"),right=Side(style="thin",color="C5CDD8"),
                                    top=Side(style="thin",color="C5CDD8"),bottom=Side(style="thin",color="C5CDD8"))
                else:
                    c.fill=fill("F8F9FA"); c.border=thin()
        ws.freeze_panes="B3"; ws.sheet_view.showGridLines=False

    # Change log sheet
    ws2 = wb.create_sheet("Change Log")
    ws2.merge_cells("A1:G1"); ws2['A1'].value="Change Log"
    ws2['A1'].font=Font(name="Arial",size=11,bold=True,color=WHITE)
    ws2['A1'].fill=fill(BLUE); ws2['A1'].alignment=Alignment(horizontal="center",vertical="center")
    ws2.row_dimensions[1].height=20
    for ci,h in enumerate(["#","Time","Type","Description","By","Resolved","Status"],1):
        c=ws2.cell(2,ci); c.value=h; c.font=Font(name="Arial",size=9,bold=True,color=WHITE)
        c.fill=fill(LBLUE); c.border=thin(); c.alignment=Alignment(horizontal="center",vertical="center")
    ws2.row_dimensions[2].height=14
    for ri,log in enumerate(change_logs):
        r=ri+3; bg=WHITE if ri%2==0 else LGRAY; ok_bg="D1FAE5" if log.get('success') else "FEE2E2"
        for ci,val in enumerate([ri+1,(log.get('created_at') or '')[:19],log.get('change_type'),
                                   log.get('description'),log.get('changed_by','system'),
                                   log.get('resolved_by',''),("OK" if log.get('success') else "FAIL")],1):
            c=ws2.cell(r,ci); c.value=val; c.font=Font(name="Arial",size=9)
            c.fill=fill(ok_bg if ci==7 else bg); c.border=thin(); c.alignment=Alignment(vertical="center")
        ws2.row_dimensions[r].height=14
    for ci,w in enumerate([4,16,16,40,18,14,10],1):
        ws2.column_dimensions[get_column_letter(ci)].width=w
    ws2.freeze_panes="A3"; ws2.sheet_view.showGridLines=False

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    fname = f"timetable_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

# ══════════════════════════════════════════
#  HEALTH + ERROR HANDLERS
# ══════════════════════════════════════════
@app.route('/api/health')
def health():
    try:
        db = get_db()
        db.command('ping')
        mongo_ok = True
    except Exception:
        mongo_ok = False
    return jsonify({"status": "ok", "mongodb": mongo_ok, "timestamp": datetime.now().isoformat()})

@app.errorhandler(404)
def not_found(e):
    if request.path.startswith('/api/'):
        return jsonify({"error": "Endpoint not found"}), 404
    return send_from_directory(FRONTEND_DIR, 'index.html')

@app.errorhandler(500)
def server_error(e):
    app.logger.error(f"500 error: {e}")
    return jsonify({"error": "Internal server error", "detail": str(e)}), 500

if __name__ == '__main__':
    port  = int(os.environ.get('PORT', 5000))
    debug = os.environ.get('DEBUG', 'false').lower() == 'true'
    host  = os.environ.get('HOST', '127.0.0.1')
    print(f"\n  Academic Timetable System (MongoDB) -> http://{host}:{port}")
    print(f"  Coordinator: admin@college.edu / admin123\n")
    app.run(host=host, port=port, debug=debug)
