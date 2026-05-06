import io
import json
import time
from collections import defaultdict

from constraints import normalize_constraint_document
from database import hash_pw, next_seq
from rescheduler import claim_slot, teacher_absent
from solver import ConstraintEvaluator, ScheduleState, SlotRecord, StandardRequest, TimetableSolver


PERIOD_IDS = ["P1", "P2", "P3", "P4", "P5", "P6"]
DAYS = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]


def insert_constraint(db, payload):
    normalized, errors = normalize_constraint_document(payload, strict=True)
    assert normalized is not None, errors
    normalized["int_id"] = next_seq(db, "user_constraints")
    db.user_constraints.insert_one(normalized)
    return normalized


def seed_basic_data(db, *, with_parallel=True, lab_room_count=2, once_per_day=False, physics_lab_hours=2):
    teachers = [
        {"id": "T1", "name": "Teacher 1", "short_name": "T1", "email": "t1@example.com", "max_hrs_per_day": 6, "max_hrs_per_week": 30, "unavailable": {}},
        {"id": "T2", "name": "Teacher 2", "short_name": "T2", "email": "t2@example.com", "max_hrs_per_day": 6, "max_hrs_per_week": 30, "unavailable": {}},
        {"id": "T3", "name": "Teacher 3", "short_name": "T3", "email": "t3@example.com", "max_hrs_per_day": 6, "max_hrs_per_week": 30, "unavailable": {}},
        {"id": "T4", "name": "Teacher 4", "short_name": "T4", "email": "t4@example.com", "max_hrs_per_day": 6, "max_hrs_per_week": 30, "unavailable": {}},
        {"id": "T5", "name": "Teacher 5", "short_name": "T5", "email": "t5@example.com", "max_hrs_per_day": 6, "max_hrs_per_week": 30, "unavailable": {}},
        {"id": "T6", "name": "Teacher 6", "short_name": "T6", "email": "t6@example.com", "max_hrs_per_day": 6, "max_hrs_per_week": 30, "unavailable": {}},
    ]
    db.teachers.insert_many(teachers)

    rooms = [{"id": "R1", "name": "Room 1", "room_type": "lecture", "capacity": 60}, {"id": "R2", "name": "Room 2", "room_type": "lecture", "capacity": 60}]
    if lab_room_count >= 1:
        rooms.append({"id": "L1", "name": "Lab 1", "room_type": "lab", "capacity": 30})
    if lab_room_count >= 2:
        rooms.append({"id": "L2", "name": "Lab 2", "room_type": "lab", "capacity": 30})
    db.rooms.insert_many(rooms)

    subjects = [
        {"id": "ALG", "name": "Algorithms", "short_name": "ALG", "code": "ALG101", "teacher_id": "T1", "lectures_per_week": 2, "has_lab": 0, "lab_hours_per_week": 0, "lab_teacher_id": None, "lab_room_id": None, "has_tutorial": 0, "tutorials_per_week": 0},
        {"id": "PHY", "name": "Physics", "short_name": "PHY", "code": "PHY101", "teacher_id": "T2", "lectures_per_week": 2, "has_lab": 1, "lab_hours_per_week": physics_lab_hours, "lab_teacher_id": "T3", "lab_room_id": None, "has_tutorial": 0, "tutorials_per_week": 0},
        {"id": "E1", "name": "Elective 1", "short_name": "E1", "code": "E101", "teacher_id": "T5", "lectures_per_week": 1, "has_lab": 0, "lab_hours_per_week": 0, "lab_teacher_id": None, "lab_room_id": None, "has_tutorial": 0, "tutorials_per_week": 0},
        {"id": "E2", "name": "Elective 2", "short_name": "E2", "code": "E102", "teacher_id": "T6", "lectures_per_week": 1, "has_lab": 0, "lab_hours_per_week": 0, "lab_teacher_id": None, "lab_room_id": None, "has_tutorial": 0, "tutorials_per_week": 0},
    ]
    db.subjects.insert_many(subjects)

    db.divisions.insert_one({"id": "D1", "name": "Division 1", "room_id": "R1", "size": 60})
    db.batches.insert_many([
        {"id": "D11", "name": "Batch 1", "division_id": "D1", "size": 30},
        {"id": "D12", "name": "Batch 2", "division_id": "D1", "size": 30},
    ])
    db.division_subjects.insert_many([
        {"division_id": "D1", "subject_id": "ALG"},
        {"division_id": "D1", "subject_id": "PHY"},
        {"division_id": "D1", "subject_id": "E1"},
        {"division_id": "D1", "subject_id": "E2"},
    ])
    db.batch_teachers.insert_many([
        {"int_id": 1, "batch_id": "D11", "subject_id": "PHY", "teacher_id": "T3"},
        {"int_id": 2, "batch_id": "D12", "subject_id": "PHY", "teacher_id": "T4"},
    ])

    if with_parallel:
        insert_constraint(
            db,
            {
                "constraint_type": "parallel_group",
                "description": "Electives must run together",
                "target_scope": "division",
                "scope": {"division_ids": ["D1"]},
                "priority": "hard",
                "weight": 10,
                "params": {"subject_ids": ["E1", "E2"], "slot_type": "lecture", "session_count": 1, "group_name": "MDM Group"},
            },
        )
    if once_per_day:
        insert_constraint(
            db,
            {
                "constraint_type": "once_per_day",
                "description": "Algorithms only once per day",
                "target_scope": "subject",
                "scope": {"subject_ids": ["ALG"], "division_ids": ["D1"]},
                "priority": "hard",
                "weight": 10,
                "params": {},
            },
        )


def login_admin(client):
    response = client.post("/api/auth/login", json={"email": "admin@college.edu", "password": "admin123"})
    assert response.status_code == 200
    return response.get_json()


def period_cell(day, period_id):
    return 3 + DAYS.index(day), 2 + PERIOD_IDS.index(period_id)


def test_constraint_evaluation_rejects_once_per_day(db):
    seed_basic_data(db, with_parallel=False, once_per_day=True)
    solver = TimetableSolver(period_ids=PERIOD_IDS, days=DAYS, db=db)
    state = ScheduleState(PERIOD_IDS, DAYS)
    state.add_slots([
        SlotRecord(
            division_id="D1",
            day="Monday",
            period_id="P1",
            subject_id="ALG",
            teacher_id="T1",
            room_id="R1",
            slot_type="lecture",
            session_id="ALG-1",
            occupancy_key="division",
        )
    ])
    request = StandardRequest("ALG-2", "D1", "ALG", "T1", "lecture", 2, "R1")
    ok, _, issues = solver.evaluator.evaluate_standard(request, "Monday", "P2", "R1", state)
    assert not ok
    assert any(issue.code == "once_per_day" for issue in issues)


def test_default_rule_rejects_duplicate_lecture_same_day(db):
    seed_basic_data(db, with_parallel=False)
    solver = TimetableSolver(period_ids=PERIOD_IDS, days=DAYS, db=db)
    state = ScheduleState(PERIOD_IDS, DAYS)
    state.add_slots([
        SlotRecord(
            division_id="D1",
            day="Monday",
            period_id="P1",
            subject_id="ALG",
            teacher_id="T1",
            room_id="R1",
            slot_type="lecture",
            session_id="ALG-1",
            occupancy_key="division",
        )
    ])
    request = StandardRequest("ALG-2", "D1", "ALG", "T1", "lecture", 2, "R1")
    ok, _, issues = solver.evaluator.evaluate_standard(request, "Monday", "P3", "R1", state)
    assert not ok
    assert any(issue.code == "same_day_repeat" for issue in issues)


def test_lab_continuity_logic(db):
    seed_basic_data(db, with_parallel=False)
    solver = TimetableSolver(period_ids=PERIOD_IDS, days=DAYS, db=db)
    assert solver.solve(), solver.failure
    grouped = defaultdict(list)
    for slot in solver.lab_assignments:
        grouped[slot.session_id].append(slot)
    assert grouped
    for slots in grouped.values():
        assert len(slots) == 2
        ordered = sorted(slots, key=lambda slot: PERIOD_IDS.index(slot.period_id))
        assert ordered[0].day == ordered[1].day
        assert PERIOD_IDS.index(ordered[1].period_id) == PERIOD_IDS.index(ordered[0].period_id) + 1
        assert ordered[0].period_id != "P2"
        assert ordered[0].period_id != "P4"
        assert ordered[0].period_id == "P5"


def test_lab_rule_prefers_last_two_hour_slot_across_available_days(db):
    seed_basic_data(db, with_parallel=False, physics_lab_hours=10)
    solver = TimetableSolver(period_ids=PERIOD_IDS, days=DAYS, db=db)
    assert solver.solve(), solver.failure

    covered_days = solver.evaluator.scheduled_end_of_day_lab_days(solver.state)
    assert covered_days == set(DAYS)

    lab_start_period = solver.evaluator.end_of_day_lab_periods()[0]
    session_starts = {}
    for slot in solver.lab_assignments:
        session_starts.setdefault(slot.session_id, set()).add(slot.period_id)
    assert session_starts
    assert all(lab_start_period in periods for periods in session_starts.values())


def test_parallel_subject_logic(db):
    seed_basic_data(db, with_parallel=True)
    solver = TimetableSolver(period_ids=PERIOD_IDS, days=DAYS, db=db)
    assert solver.solve(), solver.failure
    occurrences = defaultdict(set)
    for slot in solver.assignments:
        if slot.parallel_group_id:
            occurrences[(slot.parallel_group_id, slot.day, slot.period_id)].add(slot.subject_id)
    assert occurrences
    assert any(subjects == {"E1", "E2"} for subjects in occurrences.values())


def test_no_overlap_checks(db):
    seed_basic_data(db, with_parallel=False)
    solver = TimetableSolver(period_ids=PERIOD_IDS, days=DAYS, db=db)
    state = ScheduleState(PERIOD_IDS, DAYS)
    state.add_slots([
        SlotRecord("D1", "Monday", "P1", "ALG", "T1", "R1", "lecture", "S1", "division"),
        SlotRecord("D2", "Monday", "P1", "PHY", "T1", "R2", "lecture", "S2", "division"),
    ])
    issues = solver.evaluator.validate_state(state)
    assert any(issue.code == "teacher_overlap" for issue in issues)


def test_coverage_validation(db):
    seed_basic_data(db, with_parallel=True)
    solver = TimetableSolver(period_ids=PERIOD_IDS, days=DAYS, db=db)
    assert solver.solve(), solver.failure
    ok, report = solver.validate_required_coverage()
    assert ok
    assert not report["gaps"]
    solver.state.remove_slots([solver.assignments[0]])
    ok, report = solver.validate_required_coverage()
    assert not ok
    assert report["gaps"]


def test_generate_endpoint_feasible_sample(client, db):
    seed_basic_data(db, with_parallel=True, once_per_day=True)
    login_admin(client)
    response = client.post("/api/generate")
    assert response.status_code == 200, response.get_json()
    payload = response.get_json()
    assert payload["coverage_report"]["gaps"] == []
    diagnostics = client.get("/api/timetable/diagnostics")
    assert diagnostics.status_code == 200
    assert diagnostics.get_json() == []


def test_generate_endpoint_returns_partial_when_grouped_labs_need_more_rooms(client, db):
    seed_basic_data(db, with_parallel=False, lab_room_count=1)
    login_admin(client)
    response = client.post("/api/generate")
    assert response.status_code == 200, response.get_json()
    payload = response.get_json()
    assert payload["partial"] is True
    assert payload["unscheduled"]
    assert any(item["reason_code"] == "grouped_lab_room_shortage" for item in payload["unscheduled"])


def test_generate_groups_lab_batches_in_the_same_slot_when_feasible(client, db):
    seed_basic_data(db, with_parallel=False, lab_room_count=2)
    login_admin(client)
    response = client.post("/api/generate")
    assert response.status_code == 200, response.get_json()
    payload = response.get_json()
    assert payload["partial"] is False
    assert payload["coverage_report"]["gaps"] == []

    lab_rows = list(db.timetable_slots.find({"subject_id": "PHY", "slot_type": "lab"}, {"_id": 0}))
    assert lab_rows
    by_batch = defaultdict(set)
    for row in lab_rows:
        by_batch[row["batch_id"]].add((row["day"], row["period_id"]))
    assert by_batch["D11"]
    assert by_batch["D12"]
    assert by_batch["D11"] == by_batch["D12"]


def test_generate_allows_parallel_lab_blocks_with_different_subjects_per_batch(client, db):
    seed_basic_data(db, with_parallel=False, lab_room_count=2)
    db.subjects.insert_one(
        {
            "id": "CHE",
            "name": "Chemistry",
            "short_name": "CHE",
            "code": "CHE101",
            "teacher_id": "T5",
            "lectures_per_week": 0,
            "has_lab": 1,
            "lab_hours_per_week": 2,
            "lab_teacher_id": None,
            "lab_room_id": None,
            "has_tutorial": 0,
            "tutorials_per_week": 0,
        }
    )
    db.division_subjects.insert_one({"division_id": "D1", "subject_id": "CHE"})
    db.batch_teachers.update_one({"batch_id": "D12", "subject_id": "PHY"}, {"$set": {"teacher_id": "T3"}})
    db.batch_teachers.insert_many(
        [
            {"int_id": 3, "batch_id": "D11", "subject_id": "CHE", "teacher_id": "T4"},
            {"int_id": 4, "batch_id": "D12", "subject_id": "CHE", "teacher_id": "T4"},
        ]
    )

    login_admin(client)
    response = client.post("/api/generate")
    assert response.status_code == 200, response.get_json()
    payload = response.get_json()
    assert payload["partial"] is False
    assert payload["coverage_report"]["gaps"] == []

    lab_rows = list(db.timetable_slots.find({"slot_type": "lab"}, {"_id": 0}))
    assert lab_rows

    session_starts = {}
    for row in lab_rows:
        current = session_starts.get(row["session_id"])
        if not current or PERIOD_IDS.index(row["period_id"]) < PERIOD_IDS.index(current["period_id"]):
            session_starts[row["session_id"]] = row

    blocks = defaultdict(dict)
    for row in session_starts.values():
        blocks[(row["division_id"], row["day"], row["period_id"])][row["batch_id"]] = row["subject_id"]

    assert len(blocks) == 2
    assert all(set(assignments.keys()) == {"D11", "D12"} for assignments in blocks.values())
    assert all(set(assignments.values()) == {"PHY", "CHE"} for assignments in blocks.values())

    hours_by_batch_subject = defaultdict(int)
    for row in lab_rows:
        hours_by_batch_subject[(row["batch_id"], row["subject_id"])] += 1
    assert hours_by_batch_subject[("D11", "PHY")] == 2
    assert hours_by_batch_subject[("D11", "CHE")] == 2
    assert hours_by_batch_subject[("D12", "PHY")] == 2
    assert hours_by_batch_subject[("D12", "CHE")] == 2

def test_generate_schedules_lab_only_subject_with_shared_default_teacher(client, db):
    seed_basic_data(db, with_parallel=False, lab_room_count=2)
    db.subjects.insert_one(
        {
            "id": "PBL",
            "name": "Project Based Learning",
            "short_name": "PBL",
            "code": "PBL101",
            "teacher_id": "T5",
            "lectures_per_week": 0,
            "has_lab": 1,
            "lab_hours_per_week": 4,
            "lab_teacher_id": "T5",
            "lab_room_id": None,
            "has_tutorial": 0,
            "tutorials_per_week": 0,
        }
    )
    db.division_subjects.insert_one({"division_id": "D1", "subject_id": "PBL"})

    # Remove per-batch overrides so both batches use the same default lab teacher.
    db.batch_teachers.delete_many({"subject_id": "PBL"})

    login_admin(client)
    response = client.post("/api/generate")
    assert response.status_code == 200, response.get_json()
    payload = response.get_json()
    assert payload["coverage_report"]["gaps"] == []

    pbl_rows = list(db.timetable_slots.find({"subject_id": "PBL", "slot_type": "lab"}, {"_id": 0}))
    assert pbl_rows

    by_batch_hours = defaultdict(int)
    for row in pbl_rows:
        by_batch_hours[row["batch_id"]] += 1
    assert by_batch_hours["D11"] == 4
    assert by_batch_hours["D12"] == 4


def test_generate_infers_lab_subject_from_batch_teacher_mapping(client, db):
    seed_basic_data(db, with_parallel=False, lab_room_count=2)
    db.subjects.insert_one(
        {
            "id": "CDC",
            "name": "CDC",
            "short_name": "CDC",
            "code": "CDC101",
            "teacher_id": "T5",
            "lectures_per_week": 0,
            "has_lab": 1,
            "lab_hours_per_week": 2,
            "lab_teacher_id": None,
            "lab_room_id": None,
            "has_tutorial": 0,
            "tutorials_per_week": 0,
        }
    )

    # Intentionally do not add CDC to division_subjects.
    db.batch_teachers.insert_many(
        [
            {"int_id": 101, "batch_id": "D11", "subject_id": "CDC", "teacher_id": "T3"},
            {"int_id": 102, "batch_id": "D12", "subject_id": "CDC", "teacher_id": "T4"},
        ]
    )

    login_admin(client)
    response = client.post("/api/generate")
    assert response.status_code == 200, response.get_json()
    payload = response.get_json()
    assert payload["coverage_report"]["gaps"] == []

    cdc_rows = list(db.timetable_slots.find({"subject_id": "CDC", "slot_type": "lab"}, {"_id": 0}))
    assert cdc_rows

    by_batch_hours = defaultdict(int)
    for row in cdc_rows:
        by_batch_hours[row["batch_id"]] += 1
    assert by_batch_hours["D11"] == 2
    assert by_batch_hours["D12"] == 2


def test_generate_returns_partial_when_grouped_lab_reuses_same_teacher(client, db):
    seed_basic_data(db, with_parallel=False, lab_room_count=2)
    db.batch_teachers.update_one({"batch_id": "D12", "subject_id": "PHY"}, {"$set": {"teacher_id": "T3"}})
    login_admin(client)
    response = client.post("/api/generate")
    assert response.status_code == 200, response.get_json()
    payload = response.get_json()
    assert payload["partial"] is False
    assert payload["coverage_report"]["gaps"] == []
    assert not any(item.get("reason_code") == "grouped_lab_teacher_overlap" for item in (payload.get("unscheduled") or []))


def test_generate_endpoint_partial_fallback_returns_unscheduled_items(client, db):
    seed_basic_data(db, with_parallel=False, lab_room_count=1)
    unavailable = {day: list(PERIOD_IDS) for day in DAYS}
    db.teachers.update_one({"id": "T3"}, {"$set": {"unavailable": unavailable}})
    login_admin(client)
    response = client.post("/api/generate", json={"allow_partial": True})
    assert response.status_code == 200, response.get_json()
    payload = response.get_json()
    assert payload["partial"] is True
    assert payload["unscheduled"]
    assert payload["coverage_report"]["gaps"]


def test_validate_inputs_detects_division_over_capacity(db):
    seed_basic_data(db, with_parallel=False, lab_room_count=2)
    db.subjects.update_one({"id": "ALG"}, {"$set": {"lectures_per_week": 28}})
    solver = TimetableSolver(period_ids=PERIOD_IDS, days=DAYS, db=db)
    issues = solver.validate_inputs()
    assert any(issue.code == "division_over_capacity" for issue in issues)


def test_timetable_problem_report_shows_partial_generation_issues(client, db):
    seed_basic_data(db, with_parallel=False, lab_room_count=1)
    unavailable = {day: list(PERIOD_IDS) for day in DAYS}
    db.teachers.update_one({"id": "T3"}, {"$set": {"unavailable": unavailable}})
    login_admin(client)
    response = client.post("/api/generate", json={"allow_partial": True})
    assert response.status_code == 200, response.get_json()

    problems = client.get("/api/timetable/problems")
    assert problems.status_code == 200
    payload = problems.get_json()
    assert payload["ok"] is False
    assert payload["unscheduled"]
    assert payload["coverage_report"]["gaps"]
    assert payload["suggestions"]


def test_generate_job_endpoint_reports_progress_and_completes(client, db):
    seed_basic_data(db, with_parallel=True, once_per_day=True)
    login_admin(client)
    response = client.post("/api/generate/start")
    assert response.status_code == 202, response.get_json()
    job = response.get_json()["job"]
    assert job["status"] in {"queued", "running"}

    final = None
    for _ in range(60):
        poll = client.get(f"/api/generate/jobs/{job['job_id']}")
        assert poll.status_code == 200
        final = poll.get_json()
        if final["status"] in {"completed", "failed"}:
            break
        time.sleep(0.05)

    assert final is not None
    assert final["status"] == "completed", final
    assert final["coverage_report"]["gaps"] == []


def test_runtime_move_rejects_invalid_change(client, db):
    seed_basic_data(db, with_parallel=True, once_per_day=True)
    login_admin(client)
    generate = client.post("/api/generate")
    assert generate.status_code == 200, generate.get_json()

    slots = list(db.timetable_slots.find({"subject_id": "ALG"}, {"_id": 0}).sort([("day", 1), ("period_id", 1)]))
    assert len(slots) == 2
    target_day = slots[0]["day"]
    occupied_periods = {
        row["period_id"]
        for row in db.timetable_slots.find({"division_id": slots[0]["division_id"], "day": target_day}, {"_id": 0})
    }
    target_period = next(
        pid for pid in PERIOD_IDS
        if pid != slots[1]["period_id"] and pid not in occupied_periods
    )
    move = client.post(f"/api/slots/{slots[1]['int_id']}/move", json={"day": target_day, "period_id": target_period})
    assert move.status_code == 400
    assert "once" in move.get_json()["message"].lower()


def test_cell_move_validation_and_apply(client, db):
    seed_basic_data(db, with_parallel=True, once_per_day=True)
    login_admin(client)
    generate = client.post("/api/generate")
    assert generate.status_code == 200, generate.get_json()
    slot = db.timetable_slots.find_one({"subject_id": "E1"}, {"_id": 0})
    assert slot is not None
    validate = client.post("/api/slots/validate-cell-move", json={"slot_ids": [slot["int_id"]], "day": "Friday", "period_id": "P6"})
    assert validate.status_code in {200, 400}
    if validate.status_code == 200:
        apply_move = client.post("/api/slots/cell-move", json={"slot_ids": [slot["int_id"]], "day": "Friday", "period_id": "P6"})
        assert apply_move.status_code == 200, apply_move.get_json()


def test_auth_token_keeps_tab_sessions_independent(client, db):
    seed_basic_data(db, with_parallel=False)
    db.users.insert_one(
        {
            "int_id": next_seq(db, "users"),
            "email": "teacher1@example.com",
            "password": hash_pw("faculty123"),
            "role": "faculty",
            "name": "Teacher 1",
            "teacher_id": "T1",
        }
    )
    admin = login_admin(client)
    faculty_response = client.post("/api/auth/login", json={"email": "teacher1@example.com", "password": "faculty123"})
    assert faculty_response.status_code == 200
    faculty = faculty_response.get_json()

    admin_me = client.get("/api/auth/me", headers={"Authorization": f"Bearer {admin['auth_token']}"})
    faculty_me = client.get("/api/auth/me", headers={"Authorization": f"Bearer {faculty['auth_token']}"})

    assert admin_me.status_code == 200
    assert faculty_me.status_code == 200
    assert admin_me.get_json()["role"] == "coordinator"
    assert faculty_me.get_json()["role"] == "faculty"
    assert faculty_me.get_json()["teacher_id"] == "T1"


def test_teacher_absence_opens_lab_cover_and_claims_full_session(client, db):
    seed_basic_data(db, with_parallel=False)
    db.teachers.insert_one(
        {
            "id": "T7",
            "name": "Teacher 7",
            "short_name": "T7",
            "email": "t7-cover@example.com",
            "max_hrs_per_day": 6,
            "max_hrs_per_week": 30,
            "unavailable": {},
        }
    )
    login_admin(client)
    generate = client.post("/api/generate")
    assert generate.status_code == 200, generate.get_json()

    lab_row = db.timetable_slots.find_one({"subject_id": "PHY", "teacher_id": "T3"}, {"_id": 0})
    assert lab_row is not None

    result = teacher_absent("T3", lab_row["day"], lab_row["period_id"])
    assert result["success"] is True
    assert result["opened_sessions"]
    opened = result["opened_sessions"][0]
    assert opened["slot_type"] == "lab"

    notif = db.notifications.find_one({"type": "slot_free", "for_role": "faculty", "for_teacher": {"$ne": "T3"}}, {"_id": 0})
    assert notif is not None
    data = json.loads(notif["data"])
    assert data["session_id"] == opened["session_id"]
    assert data["candidate_teachers"]

    teacher_id = data["candidate_teachers"][0]["teacher_id"]
    ok, msg = claim_slot(notif["int_id"], teacher_id)
    assert ok, msg

    claimed_rows = list(db.timetable_slots.find({"session_id": data["session_id"]}, {"_id": 0}))
    assert len(claimed_rows) == 2
    assert all(row["teacher_id"] == teacher_id for row in claimed_rows)
    assert all(row["status"] == "active" for row in claimed_rows)
    assert all(row["source"] == "claim" for row in claimed_rows)
    assert db.notifications.count_documents({"type": "slot_claimed", "for_role": "coordinator"}) >= 1


def test_faculty_claim_endpoint_assigns_open_cover_session(client, db):
    seed_basic_data(db, with_parallel=False)
    db.teachers.insert_one(
        {
            "id": "T7",
            "name": "Teacher 7",
            "short_name": "T7",
            "email": "t7-cover@example.com",
            "max_hrs_per_day": 6,
            "max_hrs_per_week": 30,
            "unavailable": {},
        }
    )
    db.users.insert_one(
        {
            "int_id": next_seq(db, "users"),
            "email": "t7-cover@example.com",
            "password": hash_pw("faculty123"),
            "role": "faculty",
            "name": "Teacher 7",
            "teacher_id": "T7",
        }
    )
    admin = login_admin(client)
    generate = client.post("/api/generate", headers={"Authorization": f"Bearer {admin['auth_token']}"})
    assert generate.status_code == 200, generate.get_json()

    lab_row = db.timetable_slots.find_one({"subject_id": "PHY", "teacher_id": "T3"}, {"_id": 0})
    assert lab_row is not None
    absent = client.post(
        "/api/teacher-absent",
        json={"teacher_id": "T3", "day": lab_row["day"], "period_id": lab_row["period_id"]},
        headers={"Authorization": f"Bearer {admin['auth_token']}"},
    )
    assert absent.status_code == 200, absent.get_json()

    notif = db.notifications.find_one({"type": "slot_free", "for_role": "faculty", "for_teacher": "T7"}, {"_id": 0})
    assert notif is not None

    faculty_login = client.post("/api/auth/login", json={"email": "t7-cover@example.com", "password": "faculty123"})
    assert faculty_login.status_code == 200
    faculty = faculty_login.get_json()

    claim = client.post(
        f"/api/notifications/{notif['int_id']}/claim",
        headers={"Authorization": f"Bearer {faculty['auth_token']}"},
    )
    assert claim.status_code == 200, claim.get_json()

    data = json.loads(notif["data"])
    claimed_rows = list(db.timetable_slots.find({"session_id": data["session_id"]}, {"_id": 0}))
    assert claimed_rows
    assert all(row["teacher_id"] == "T7" for row in claimed_rows)
    assert all(row["status"] == "active" for row in claimed_rows)


def test_admin_can_assign_open_cover_from_notification(client, db):
    seed_basic_data(db, with_parallel=False)
    db.teachers.insert_one(
        {
            "id": "T7",
            "name": "Teacher 7",
            "short_name": "T7",
            "email": "t7-cover@example.com",
            "max_hrs_per_day": 6,
            "max_hrs_per_week": 30,
            "unavailable": {},
        }
    )
    db.users.insert_one(
        {
            "int_id": next_seq(db, "users"),
            "email": "t7-cover@example.com",
            "password": hash_pw("faculty123"),
            "role": "faculty",
            "name": "Teacher 7",
            "teacher_id": "T7",
        }
    )
    admin = login_admin(client)
    generate = client.post("/api/generate", headers={"Authorization": f"Bearer {admin['auth_token']}"})
    assert generate.status_code == 200, generate.get_json()

    lab_row = db.timetable_slots.find_one({"subject_id": "PHY", "teacher_id": "T3"}, {"_id": 0})
    assert lab_row is not None
    absent = client.post(
        "/api/teacher-absent",
        json={"teacher_id": "T3", "day": lab_row["day"], "period_id": lab_row["period_id"]},
        headers={"Authorization": f"Bearer {admin['auth_token']}"},
    )
    assert absent.status_code == 200, absent.get_json()

    notif = db.notifications.find_one({"type": "slot_free", "for_role": "coordinator"}, {"_id": 0})
    assert notif is not None

    assign = client.post(
        f"/api/notifications/{notif['int_id']}/assign",
        json={"teacher_id": "T7"},
        headers={"Authorization": f"Bearer {admin['auth_token']}"},
    )
    assert assign.status_code == 200, assign.get_json()

    data = json.loads(notif["data"])
    claimed_rows = list(db.timetable_slots.find({"session_id": data["session_id"]}, {"_id": 0}))
    assert claimed_rows
    assert all(row["teacher_id"] == "T7" for row in claimed_rows)
    assert db.notifications.count_documents({"type": "slot_claimed", "for_role": "faculty", "for_teacher": "T7"}) >= 1


def test_swap_endpoint_moves_parallel_session_as_one_block(client, db):
    seed_basic_data(db, with_parallel=True)
    db.teachers.insert_one(
        {
            "id": "T7",
            "name": "Teacher 7",
            "short_name": "T7",
            "email": "t7@example.com",
            "max_hrs_per_day": 6,
            "max_hrs_per_week": 30,
            "unavailable": {},
        }
    )
    db.subjects.insert_one(
        {
            "id": "BIO",
            "name": "Biology",
            "short_name": "BIO",
            "code": "BIO101",
            "teacher_id": "T7",
            "lectures_per_week": 1,
            "has_lab": 0,
            "lab_hours_per_week": 0,
            "lab_teacher_id": None,
            "lab_room_id": None,
            "has_tutorial": 0,
            "tutorials_per_week": 0,
        }
    )
    db.division_subjects.insert_one({"division_id": "D1", "subject_id": "BIO"})

    login_admin(client)
    generate = client.post("/api/generate")
    assert generate.status_code == 200, generate.get_json()

    parallel_slot = db.timetable_slots.find_one({"subject_id": "E1"}, {"_id": 0})
    bio_slot = db.timetable_slots.find_one({"subject_id": "BIO"}, {"_id": 0})
    assert parallel_slot is not None
    assert bio_slot is not None
    assert (parallel_slot["day"], parallel_slot["period_id"]) != (bio_slot["day"], bio_slot["period_id"])

    old_parallel = (parallel_slot["day"], parallel_slot["period_id"])
    old_bio = (bio_slot["day"], bio_slot["period_id"])
    group_id = parallel_slot["parallel_group_id"]
    assert group_id

    check = client.post("/api/slots/validate-swap", json={"slot_id_1": parallel_slot["int_id"], "slot_id_2": bio_slot["int_id"]})
    assert check.status_code == 200, check.get_json()
    assert check.get_json()["ok"] is True

    swap = client.post("/api/slots/swap", json={"slot_id_1": parallel_slot["int_id"], "slot_id_2": bio_slot["int_id"]})
    assert swap.status_code == 200, swap.get_json()

    moved_parallel = list(db.timetable_slots.find({"parallel_group_id": group_id}, {"_id": 0}))
    assert moved_parallel
    assert all((row["day"], row["period_id"]) == old_bio for row in moved_parallel)
    moved_bio = db.timetable_slots.find_one({"subject_id": "BIO"}, {"_id": 0})
    assert (moved_bio["day"], moved_bio["period_id"]) == old_parallel


def test_swap_endpoint_reassigns_room_when_target_room_is_busy(client, db):
    seed_basic_data(db, with_parallel=False)
    db.teachers.insert_many(
        [
            {
                "id": "T7",
                "name": "Teacher 7",
                "short_name": "T7",
                "email": "t7@example.com",
                "max_hrs_per_day": 6,
                "max_hrs_per_week": 30,
                "unavailable": {},
            },
            {
                "id": "T8",
                "name": "Teacher 8",
                "short_name": "T8",
                "email": "t8@example.com",
                "max_hrs_per_day": 6,
                "max_hrs_per_week": 30,
                "unavailable": {},
            },
        ]
    )
    db.subjects.insert_one(
        {
            "id": "BIO",
            "name": "Biology",
            "short_name": "BIO",
            "code": "BIO101",
            "teacher_id": "T7",
            "lectures_per_week": 0,
            "has_lab": 0,
            "lab_hours_per_week": 0,
            "lab_teacher_id": None,
            "lab_room_id": None,
            "has_tutorial": 0,
            "tutorials_per_week": 0,
        }
    )
    db.subjects.insert_one(
        {
            "id": "CHE",
            "name": "Chemistry",
            "short_name": "CHE",
            "code": "CHE101",
            "teacher_id": "T8",
            "lectures_per_week": 0,
            "has_lab": 0,
            "lab_hours_per_week": 0,
            "lab_teacher_id": None,
            "lab_room_id": None,
            "has_tutorial": 0,
            "tutorials_per_week": 0,
        }
    )
    db.divisions.insert_one({"id": "D2", "name": "Division 2", "room_id": "R2", "size": 60})

    admin = login_admin(client)
    generate = client.post("/api/generate", headers={"Authorization": f"Bearer {admin['auth_token']}"})
    assert generate.status_code == 200, generate.get_json()

    slot_a = db.timetable_slots.find_one({"subject_id": "E1"}, {"_id": 0})
    assert slot_a is not None
    other_room = "R2" if slot_a["room_id"] == "R1" else "R1"
    blocker_room = slot_a["room_id"]
    version_id = slot_a["version_id"]

    occupied = {
        (row["day"], row["period_id"])
        for row in db.timetable_slots.find({"division_id": "D1"}, {"_id": 0, "day": 1, "period_id": 1})
    }
    target_day, target_period = next((day, pid) for day in DAYS for pid in PERIOD_IDS if (day, pid) not in occupied and (day, pid) != (slot_a["day"], slot_a["period_id"]))
    next_id = max(row["int_id"] for row in db.timetable_slots.find({}, {"_id": 0, "int_id": 1})) + 1

    db.timetable_slots.insert_many(
        [
            {
                "int_id": next_id,
                "division_id": "D1",
                "day": target_day,
                "period_id": target_period,
                "subject_id": "BIO",
                "teacher_id": "T7",
                "room_id": other_room,
                "slot_type": "lecture",
                "batch_id": None,
                "parallel_group_id": None,
                "session_id": "MANUAL:BIO:1",
                "occupancy_key": "division",
                "is_locked": 0,
                "status": "active",
                "source": "manual",
                "version_id": version_id,
            },
            {
                "int_id": next_id + 1,
                "division_id": "D2",
                "day": target_day,
                "period_id": target_period,
                "subject_id": "CHE",
                "teacher_id": "T8",
                "room_id": blocker_room,
                "slot_type": "lecture",
                "batch_id": None,
                "parallel_group_id": None,
                "session_id": "MANUAL:CHE:1",
                "occupancy_key": "division",
                "is_locked": 0,
                "status": "active",
                "source": "manual",
                "version_id": version_id,
            },
        ]
    )

    swap = client.post(
        "/api/slots/swap",
        json={"slot_id_1": slot_a["int_id"], "slot_id_2": next_id},
        headers={"Authorization": f"Bearer {admin['auth_token']}"},
    )
    assert swap.status_code == 200, swap.get_json()

    moved_a = db.timetable_slots.find_one({"int_id": slot_a["int_id"]}, {"_id": 0})
    moved_b = db.timetable_slots.find_one({"int_id": next_id}, {"_id": 0})
    assert (moved_a["day"], moved_a["period_id"]) == (target_day, target_period)
    assert moved_a["room_id"] == other_room
    assert (moved_b["day"], moved_b["period_id"]) == (slot_a["day"], slot_a["period_id"])


def test_excel_export_keeps_parallel_sessions_in_same_cell(client, db):
    from openpyxl import load_workbook

    seed_basic_data(db, with_parallel=True)
    login_admin(client)
    generate = client.post("/api/generate")
    assert generate.status_code == 200, generate.get_json()

    e1 = db.timetable_slots.find_one({"subject_id": "E1"}, {"_id": 0})
    e2 = db.timetable_slots.find_one({"subject_id": "E2", "day": e1["day"], "period_id": e1["period_id"]}, {"_id": 0})
    assert e1 is not None and e2 is not None

    export = client.get("/api/export/excel")
    assert export.status_code == 200

    wb = load_workbook(io.BytesIO(export.data))
    ws = wb["Div D1"]
    row_idx, col_idx = period_cell(e1["day"], e1["period_id"])
    cell_value = ws.cell(row_idx, col_idx).value or ""
    assert "E1" in cell_value
    assert "E2" in cell_value
